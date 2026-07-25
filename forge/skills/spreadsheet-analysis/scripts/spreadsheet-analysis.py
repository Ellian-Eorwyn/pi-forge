#!/usr/bin/env python3

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import os
import re
import statistics
import sys
from collections import Counter
from copy import copy
from datetime import date, datetime, time, timezone
from pathlib import Path

# Shared forge embeddings client lives at forge/lib; this script is at
# forge/skills/spreadsheet-analysis/scripts/spreadsheet-analysis.py.
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "lib"))
import forge_embeddings
import forge_llm
import forge_verify
import run_state


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx"}
RUN_SCHEMA_VERSION = 1
CLUSTER_SCHEMA_VERSION = 1
RESULT_STATUSES = {"completed", "failed", "skipped", "needs_review"}

# Default cosine similarity for grouping rows by an embedded column. Raise it
# (~0.92+) for tight duplicate detection; lower it (~0.6-0.75) for broader
# topical categorization.
DEFAULT_CLUSTER_THRESHOLD = 0.85

# Maximum characters of key text embedded per row.
CLUSTER_KEY_CHARS = 2000

# Enrichment reuses one answer across rows that mean the same thing. The
# organizer's 0.97 near-duplicate threshold does not transfer here: it asks "is
# this the same text twice", while reuse asks "do these rows deserve the same
# answer", which paraphrases satisfy at a looser bound. Measured on a support
# ticket sheet, three genuine duplicate groups sat between 0.945 and 0.964, so
# 0.97 reused nothing at all; 0.85 through 0.92 found exactly those three groups
# with no false merge, and 0.80 began merging distinct problems. 0.92 is the top
# of that plateau — the most conservative setting that still works.
ROW_REUSE_THRESHOLD = 0.92
# A spreadsheet cell holding more than this is a report, not a value.
MAX_ENRICHMENT_CHARACTERS = 2000
DEFAULT_VERIFY_SAMPLE = 25
# When review rejects a merge, its members are re-answered individually on the
# thinking model. Past this many, that costs more than it is worth and the rest
# are handed to a human instead of being answered badly or answered slowly.
MAX_REBUILT_ROWS = 50


def fail(message, exit_code=1):
    print(f"Error: {message}", file=sys.stderr)
    raise SystemExit(exit_code)


def utc_now():
    return datetime.now(timezone.utc).isoformat()


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_value(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return str(value)
        return value
    return str(value)


def display_header(value):
    if value is None:
        return ""
    return str(value)


def blank(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def require_source(raw_path):
    path = Path(raw_path).expanduser().resolve()
    if not path.exists():
        fail(f"input does not exist: {path}")
    if not path.is_file():
        fail(f"input is not a file: {path}")
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        fail(f"unsupported input format {path.suffix or '(none)'}; expected .csv, .tsv, or .xlsx")
    return path


def require_new_directory(raw_path):
    path = Path(raw_path).expanduser().resolve()
    if path.exists():
        fail(f"output already exists: {path}")
    path.mkdir(parents=True)
    return path


def require_run_directory(raw_path):
    path = Path(raw_path).expanduser().resolve()
    if not path.is_dir():
        fail(f"run directory does not exist: {path}")
    if not (path / "run.json").is_file():
        fail(f"run.json is missing: {path}")
    return path


def openpyxl_module():
    try:
        import openpyxl
    except ImportError:
        fail("XLSX support requires openpyxl; install it for the active Python 3 environment")
    return openpyxl


def csv_rows(path):
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.reader(handle, delimiter=delimiter))
    except UnicodeDecodeError:
        fail(f"input is not valid UTF-8: {path}")


def load_tables(path, requested_sheet=None, all_xlsx_sheets=False):
    extension = path.suffix.lower()
    if extension in {".csv", ".tsv"}:
        if requested_sheet:
            fail("--sheet is only valid for XLSX input")
        rows = csv_rows(path)
        return [{"name": "Data", "rows": rows}]

    openpyxl = openpyxl_module()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_links=True)
    try:
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                fail(f"sheet not found: {requested_sheet}")
            worksheets = [workbook[requested_sheet]]
        elif all_xlsx_sheets:
            worksheets = list(workbook.worksheets)
        else:
            worksheets = [workbook.active]
        tables = []
        for worksheet in worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            tables.append({"name": worksheet.title, "rows": rows})
        return tables
    finally:
        workbook.close()


def normalized_width(rows):
    return max((len(row) for row in rows), default=0)


def pad_rows(rows, width):
    return [row + [None] * (width - len(row)) for row in rows]


def infer_type(values):
    nonblank = [value for value in values if not blank(value)]
    if not nonblank:
        return "empty"
    kinds = set()
    for value in nonblank:
        if isinstance(value, bool):
            kinds.add("boolean")
        elif isinstance(value, int):
            kinds.add("integer")
        elif isinstance(value, float):
            kinds.add("number")
        elif isinstance(value, (datetime, date, time)):
            kinds.add("date")
        elif isinstance(value, str):
            kinds.add(infer_string_type(value))
        else:
            kinds.add("text")
    if kinds <= {"integer"}:
        return "integer"
    if kinds <= {"integer", "number"}:
        return "number"
    if len(kinds) == 1:
        return next(iter(kinds))
    return "mixed"


def infer_string_type(value):
    stripped = value.strip()
    if stripped.startswith("="):
        return "formula"
    if stripped.lower() in {"true", "false"}:
        return "boolean"
    if re.fullmatch(r"[+-]?(?:0|[1-9]\d*)", stripped):
        unsigned = stripped.lstrip("+-")
        if len(unsigned) == 1 or not unsigned.startswith("0"):
            return "integer"
    if re.fullmatch(r"[+-]?(?:(?:0|[1-9]\d*)\.\d+|(?:0|[1-9]\d*)[eE][+-]?\d+|(?:0|[1-9]\d*)\.\d+[eE][+-]?\d+)", stripped):
        try:
            if math.isfinite(float(stripped)):
                return "number"
        except ValueError:
            pass
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[T ][^\s]+)?", stripped):
        try:
            datetime.fromisoformat(stripped.replace("Z", "+00:00"))
            return "date"
        except ValueError:
            pass
    return "text"


def counter_key(value):
    value = json_value(value)
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def counter_label(key):
    return json.loads(key)


def numeric_distribution(values, inferred):
    if inferred not in {"integer", "number"}:
        return None, []
    numeric = []
    for value in values:
        if blank(value) or isinstance(value, bool):
            continue
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(number):
            numeric.append(number)
    if not numeric:
        return None, []
    ordered = sorted(numeric)
    distribution = {
        "minimum": min(ordered),
        "maximum": max(ordered),
        "mean": statistics.fmean(ordered),
        "median": statistics.median(ordered),
    }
    if len(ordered) < 4:
        return distribution, []
    quartiles = statistics.quantiles(ordered, n=4, method="inclusive")
    lower, upper = quartiles[0], quartiles[2]
    spread = upper - lower
    low_fence = lower - 1.5 * spread
    high_fence = upper + 1.5 * spread
    unusual = [value for value in ordered if value < low_fence or value > high_fence][:10]
    return distribution, unusual


def profile_table(table, max_categories):
    rows = table["rows"]
    width = normalized_width(rows)
    if not rows or width == 0:
        return {
            "name": table["name"],
            "rowCount": 0,
            "dataRowCount": 0,
            "columnCount": 0,
            "duplicateDataRows": 0,
            "columns": [],
            "warnings": ["Sheet or table is empty."],
        }
    padded = pad_rows(rows, width)
    headers = [display_header(value) for value in padded[0]]
    data = padded[1:]
    warnings = []
    if any(header == "" for header in headers):
        warnings.append("One or more header cells are blank.")
    duplicate_headers = sorted(header for header, count in Counter(headers).items() if header and count > 1)
    if duplicate_headers:
        warnings.append(f"Duplicate headers: {', '.join(duplicate_headers)}")
    row_keys = [tuple(counter_key(value) for value in row) for row in data if not all(blank(value) for value in row)]
    duplicate_rows = sum(count - 1 for count in Counter(row_keys).values() if count > 1)
    columns = []
    for index, header in enumerate(headers):
        values = [row[index] for row in data]
        nonblank = [value for value in values if not blank(value)]
        frequencies = Counter(counter_key(value) for value in nonblank)
        top_values = [
            {"value": counter_label(key), "count": count}
            for key, count in frequencies.most_common(max_categories)
        ]
        inferred = infer_type(values)
        numeric, numeric_unusual = numeric_distribution(values, inferred)
        unusual = numeric_unusual
        if numeric is None and len(nonblank) >= 10 and len(frequencies) <= len(nonblank) / 2:
            unusual = [
                counter_label(key)
                for key, count in frequencies.items()
                if count / len(nonblank) <= 0.05
            ][:10]
        columns.append(
            {
                "index": index + 1,
                "header": header,
                "inferredType": inferred,
                "missingCount": len(values) - len(nonblank),
                "nonMissingCount": len(nonblank),
                "uniqueCount": len(frequencies),
                "topValues": top_values,
                "numericDistribution": numeric,
                "unusualValues": unusual,
            }
        )
    return {
        "name": table["name"],
        "rowCount": len(rows),
        "dataRowCount": len(data),
        "columnCount": width,
        "duplicateDataRows": duplicate_rows,
        "columns": columns,
        "warnings": warnings,
    }


def markdown_value(value):
    if value is None:
        return ""
    text = str(value).replace("|", "\\|").replace("\n", " ")
    return text if len(text) <= 80 else f"{text[:77]}..."


def profile_markdown(profile):
    source = profile["source"]
    lines = [
        "# Data Profile",
        "",
        "## Source and Provenance",
        "",
        f"- Path: `{source['path']}`",
        f"- Format: `{source['format']}`",
        f"- SHA-256: `{source['sha256']}`",
        f"- Size: {source['sizeBytes']} bytes",
        f"- Generated: {profile['generatedAt']}",
        "",
    ]
    for sheet in profile["sheets"]:
        lines.extend(
            [
                f"## Sheet: {sheet['name']}",
                "",
                f"- Data rows: {sheet['dataRowCount']}",
                f"- Columns: {sheet['columnCount']}",
                f"- Duplicate data rows beyond first occurrence: {sheet['duplicateDataRows']}",
            ]
        )
        for warning in sheet["warnings"]:
            lines.append(f"- Warning: {warning}")
        lines.extend(
            [
                "",
                "| # | Header | Inferred type | Missing | Unique | Common values | Unusual values |",
                "|---:|---|---|---:|---:|---|---|",
            ]
        )
        for column in sheet["columns"]:
            common = ", ".join(
                f"{markdown_value(item['value'])} ({item['count']})" for item in column["topValues"][:5]
            )
            unusual = ", ".join(markdown_value(value) for value in column["unusualValues"][:5])
            lines.append(
                f"| {column['index']} | {markdown_value(column['header'])} | {column['inferredType']} | "
                f"{column['missingCount']} | {column['uniqueCount']} | {common} | {unusual} |"
            )
        lines.append("")
    lines.extend(
        [
            "## Interpretation Limits",
            "",
            "- Types and unusual values are heuristic and require review before transformation.",
            "- Empty strings and null cells are counted as missing; other missing-value tokens are not assumed.",
            "- Profiles are bounded and do not replace domain-specific validation.",
            "",
        ]
    )
    return "\n".join(lines)


def command_doctor(args):
    available = importlib.util.find_spec("openpyxl") is not None
    version = None
    if available:
        import openpyxl
        version = openpyxl.__version__
    embeddings = forge_embeddings.embeddings_doctor()
    result = {
        "python": sys.version.split()[0],
        "csvTsv": True,
        "xlsx": available,
        "openpyxlVersion": version,
        "embeddings": embeddings,
        "remediation": None if available else "Install openpyxl for the active Python 3 environment.",
    }
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(f"Python: {result['python']}")
        print("CSV/TSV: available")
        print(f"XLSX: {'available via openpyxl ' + version if available else 'unavailable'}")
        reach = "reachable" if embeddings["reachable"] else "unreachable"
        print(f"Embeddings ({embeddings['url']}): {reach} - {embeddings['detail']}")
        print("  Required by the 'cluster' command for fuzzy record linkage and semantic grouping.")
        if result["remediation"]:
            print(f"Action: {result['remediation']}")


def command_inspect(args):
    source = require_source(args.input)
    output = require_new_directory(args.output)
    try:
        tables = load_tables(source, requested_sheet=args.sheet, all_xlsx_sheets=args.sheet is None)
        profile = {
            "schemaVersion": 1,
            "generatedAt": utc_now(),
            "source": {
                "path": str(source),
                "format": source.suffix.lower().lstrip("."),
                "sha256": sha256(source),
                "sizeBytes": source.stat().st_size,
            },
            "sheets": [profile_table(table, args.max_categories) for table in tables],
        }
        (output / "data_profile.json").write_text(json.dumps(profile, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        (output / "data_profile.md").write_text(profile_markdown(profile), encoding="utf-8")
        (output / "transform_log.md").write_text(
            "# Transformation Log\n\nNo data transformations were performed. This run created a bounded profile only.\n",
            encoding="utf-8",
        )
    except BaseException:
        try:
            output.rmdir()
        except OSError:
            pass
        raise
    print(json.dumps({"source": str(source), "output": str(output), "sheets": [table["name"] for table in tables]}))


def validate_headers(headers):
    if not headers:
        fail("the selected table has no header row")
    if any(header == "" for header in headers):
        fail("row enrichment requires every header cell to be nonblank")
    duplicates = sorted(header for header, count in Counter(headers).items() if count > 1)
    if duplicates:
        fail(f"row enrichment requires unique headers; duplicates: {', '.join(duplicates)}")


def command_row_init(args):
    source = require_source(args.input)
    output = Path(args.output).expanduser().resolve()
    tables = load_tables(source, requested_sheet=args.sheet, all_xlsx_sheets=False)
    table = tables[0]
    rows = table["rows"]
    width = normalized_width(rows)
    if not rows or width == 0:
        fail("the selected table is empty")
    padded = pad_rows(rows, width)
    headers = [display_header(value) for value in padded[0]]
    validate_headers(headers)
    if args.column in headers:
        fail(f"output column already exists: {args.column}")
    selected = args.input_columns if args.input_columns else headers
    unknown = [header for header in selected if header not in headers]
    if unknown:
        fail(f"input columns not found: {', '.join(unknown)}")
    if len(set(selected)) != len(selected):
        fail("--input-columns contains duplicates")
    start_row = args.start_row if args.start_row is not None else 2
    end_row = args.end_row if args.end_row is not None else len(padded)
    if start_row < 2:
        fail("--start-row must be 2 or greater because row 1 is the header")
    if end_row < start_row:
        fail("--end-row must be greater than or equal to --start-row")
    end_row = min(end_row, len(padded))
    header_indexes = {header: index for index, header in enumerate(headers)}
    eligible = []
    blank_rows = []
    row_data = {}
    for row_number in range(start_row, end_row + 1):
        row = padded[row_number - 1]
        if all(blank(value) for value in row):
            blank_rows.append(row_number)
            continue
        eligible.append(row_number)
        row_data[str(row_number)] = {header: json_value(row[header_indexes[header]]) for header in selected}
    source_info = {
        "path": str(source),
        "basename": source.name,
        "format": source.suffix.lower().lstrip("."),
        "sha256": sha256(source),
        "sizeBytes": source.stat().st_size,
    }
    instruction = (args.instruction or "").strip()
    options = {
        "sheet": table["name"],
        "column": args.column,
        "inputColumns": selected,
        "startRow": start_row,
        "endRow": end_row,
    }
    # Only fingerprinted when supplied, so a run created without an instruction
    # still resumes unchanged. When there is one it belongs in the fingerprint:
    # it defines the output as much as the column selection does.
    if instruction:
        options["instruction"] = instruction
    configuration = {
        "workflow": "spreadsheet-row-enrichment",
        "command": "row-init",
        "input": {"path": str(source)},
        "options": options,
    }
    if output.exists():
        try:
            state = run_state.load_run_state(output, "spreadsheet-row-enrichment")
            run_state.assert_compatible_run(state, configuration)
        except (OSError, ValueError) as error:
            fail(str(error))
        print(json.dumps({"runDirectory": str(output), "resumed": True, "status": state["status"], "phase": state["phase"], "nextAction": state.get("nextAction")}))
        return
    output.mkdir(parents=True)
    run = {
        "schemaVersion": RUN_SCHEMA_VERSION,
        "createdAt": utc_now(),
        "source": source_info,
        "sheet": table["name"],
        "headerRow": 1,
        "outputColumn": args.column,
        "inputColumns": selected,
        "instruction": instruction,
        "startRow": start_row,
        "endRow": end_row,
        "eligibleRows": eligible,
        "blankRows": blank_rows,
        "rows": row_data,
    }
    run_state.atomic_write_json(output / "run.json", run)
    run_state.atomic_write_json(output / "source_manifest.json", source_info)
    run_state.atomic_write_text(output / "row_results.jsonl", "")
    state = run_state.create_run_state(
        "spreadsheet-row-enrichment",
        "row-init",
        configuration["input"],
        configuration["options"],
        items=[{"id": row_id, "status": "pending", "attempts": 0, "transient": False} for row_id in eligible],
        phase="enriching",
        next_action="row-next",
    )
    run_state.initialize_run_state(output, state)
    print(json.dumps({"runDirectory": str(output), "resumed": False, "sheet": table["name"], "eligibleRows": len(eligible), "blankRows": len(blank_rows)}))


def load_run(run_directory):
    try:
        run = json.loads((run_directory / "run.json").read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        fail(f"could not read run.json: {error}")
    if run.get("schemaVersion") != RUN_SCHEMA_VERSION:
        fail(f"unsupported run schema version: {run.get('schemaVersion')}")
    return run


def load_results(run_directory, strict=True):
    path = run_directory / "row_results.jsonl"
    if not path.is_file():
        fail(f"row_results.jsonl is missing: {path}")
    try:
        results, warnings = run_state.read_jsonl_recover_tail(path, repair=True)
    except ValueError as error:
        fail(str(error))
    for warning in warnings:
        print(f"Warning: {warning}", file=sys.stderr)
    seen = set()
    for result in results:
        row_id = result.get("rowId")
        if strict and row_id in seen:
            fail(f"duplicate result for row {row_id}")
        seen.add(row_id)
    return results


def next_pending(run, results):
    completed = {result.get("rowId") for result in results}
    for row_id in run["eligibleRows"]:
        if row_id not in completed:
            return row_id
    return None


def command_row_next(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    results = load_results(run_directory)
    row_id = next_pending(run, results)
    if row_id is None:
        print(json.dumps({"complete": True, "processed": len(results), "total": len(run["eligibleRows"])}))
        return
    print(
        json.dumps(
            {
                "complete": False,
                "rowId": row_id,
                "sourceRow": row_id,
                "sheet": run["sheet"],
                "input": run["rows"][str(row_id)],
                "outputColumn": run["outputColumn"],
                "instruction": run.get("instruction", ""),
                "progress": {"processed": len(results), "total": len(run["eligibleRows"])},
            },
            ensure_ascii=False,
        )
    )


def command_row_status(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    try:
        state = run_state.load_run_state(run_directory, "spreadsheet-row-enrichment")
    except ValueError as error:
        fail(str(error))
    results = load_results(run_directory)
    source = Path(run["source"]["path"])
    changed = not source.is_file() or sha256(source) != run["source"]["sha256"]
    print(json.dumps({"runDirectory": str(run_directory), "status": state["status"], "phase": state["phase"], "nextAction": state.get("nextAction"), "processed": len(results), "total": len(run["eligibleRows"]), "inputChanged": changed, "refreshRequired": changed}, indent=2))


def command_row_retry(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    results = load_results(run_directory)
    targets = {
        result["rowId"]
        for result in results
        if result["status"] == "failed" and (args.all_failed or result["rowId"] == args.item)
    }
    if not targets:
        fail(f"failed row not found: {args.item}" if args.item else "run has no failed rows")
    retained = [result for result in results if result["rowId"] not in targets]
    retained.sort(key=lambda value: run["eligibleRows"].index(value["rowId"]))
    run_state.atomic_write_text(run_directory / "row_results.jsonl", "" if not retained else "\n".join(json.dumps(value, ensure_ascii=False) for value in retained) + "\n")
    def retried(state):
        for item in state["items"]:
            if item["id"] in targets:
                item.update({"status": "pending", "attempts": 0, "error": None, "transient": False})
        state["status"] = "running"
        state["phase"] = "enriching"
        state["nextAction"] = "row-next"
        return state
    run_state.update_run_state(run_directory, retried, {"type": "items_retried", "itemIds": sorted(targets)})
    print(json.dumps({"runDirectory": str(run_directory), "retried": len(targets), "nextAction": "row-next"}))


def command_row_record(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    results = load_results(run_directory)
    expected = next_pending(run, results)
    if expected is None:
        fail("the run is already complete")
    if args.row_id != expected:
        fail(f"rows must be recorded sequentially; expected row {expected}, received {args.row_id}")
    if args.status == "completed":
        if not args.value_file:
            fail("completed results require --value-file")
        value_path = Path(args.value_file).expanduser().resolve()
        if not value_path.is_file():
            fail(f"value file does not exist: {value_path}")
        try:
            value = value_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            fail(f"value file is not valid UTF-8: {value_path}")
        if not value.strip():
            fail("completed results require a nonblank value; use an explicit non-completed status instead")
    else:
        if args.value_file:
            fail("--value-file is only valid with --status completed")
        if not args.note:
            fail(f"--status {args.status} requires --note")
        value = None
    remaining = record_row(run_directory, run, args.row_id, args.status, value, args.note)
    print(json.dumps({"recorded": args.row_id, "status": args.status, "remaining": remaining}))


def record_row(run_directory, run, row_id, status, value, note, provenance=None):
    """Commit one row's disposition. Shared by the CLI and the worker so both go
    through the same ordering and state transition.

    The file is always rewritten in eligible-row order, which ``validate``
    requires, so re-recording a row replaces it rather than duplicating it.
    """
    results = load_results(run_directory)
    result = {
        "rowId": row_id,
        "status": status,
        "value": value,
        "note": note,
        "recordedAt": utc_now(),
    }
    if provenance:
        result.update(provenance)
    combined = [entry for entry in results if entry["rowId"] != row_id]
    combined.append(result)
    ordered = sorted(combined, key=lambda entry: run["eligibleRows"].index(entry["rowId"]))
    run_state.atomic_write_text(
        run_directory / "row_results.jsonl",
        "\n".join(json.dumps(entry, ensure_ascii=False) for entry in ordered) + "\n",
    )

    def recorded(state):
        for item in state["items"]:
            if item["id"] == row_id:
                item.update({"status": status, "attempts": item.get("attempts", 0) + 1, "error": note if status == "failed" else None})
        if all(item["status"] != "pending" for item in state["items"]):
            state["phase"] = "finalizing"
            state["nextAction"] = "row-finalize"
        return state

    run_state.update_run_state(run_directory, recorded, {"type": "item_recorded", "itemId": row_id, "status": status})
    return len(run["eligibleRows"]) - len(ordered)


# --------------------------------------------------------------------------- #
# Stateless enrichment worker
# --------------------------------------------------------------------------- #

# One row per call, with no conversation carried between them. The instruction
# and the column contract live here so the prefix stays byte-stable across every
# call in a run; only the row travels in the user message.
ENRICHMENT_SYSTEM = """You fill in one spreadsheet column, one row at a time.

Output column: "{column}"
Input columns available on each row: {inputs}

What the output column must contain:
{instruction}

Return exactly one JSON object and nothing else, either:
{{"status": "completed", "value": "<the cell value>"}}
or:
{{"status": "needs_review", "note": "<what is missing or ambiguous>"}}

Rules:
- The value is one spreadsheet cell. Keep it brief.
- Answer every row in the same shape: the same kind of answer, the same wording
  pattern, the same units, the same capitalization. A column whose cells are
  each phrased differently cannot be sorted, filtered, or counted, which is what
  a column is for. Decide the form from the instruction, then hold it.
- Use only what the row provides. Never invent an identifier, quantity, date, or
  name the row does not contain.
- When the row does not support an answer, return needs_review with a note
  rather than guessing or writing a placeholder.
- Return no explanation and no text outside the JSON object.
"""


def progress(message):
    """Per-row progress on stderr; stdout stays one JSON result."""
    print(message, file=sys.stderr, flush=True)


class UserError(RuntimeError):
    """A row could not be enriched; the run records it and continues."""


def enrichment_system_prompt(run, instruction):
    return ENRICHMENT_SYSTEM.format(
        column=run["outputColumn"],
        inputs=", ".join(f'"{column}"' for column in run["inputColumns"]),
        instruction=instruction,
    )


def row_text(run, row_id):
    row = run["rows"][str(row_id)]
    parts = [f"{column}: {row.get(column)}" for column in run["inputColumns"] if not blank(row.get(column))]
    return " | ".join(parts)[:CLUSTER_KEY_CHARS]


def reuse_clusters(run, row_ids, args):
    """Group rows that mean the same thing, so one answer can serve all of them.

    Returns ``(cluster_of_row, note)``. Embeddings are an optimization here
    rather than a requirement: if the endpoint is unavailable every row becomes
    its own cluster and the run costs one call per row, which is what it cost
    before this existed. That degradation is reported, never silent.
    """
    singletons = {row_id: index for index, row_id in enumerate(row_ids)}
    if args.no_cluster:
        return singletons, "row reuse disabled with --no-cluster"
    if len(row_ids) < 2:
        return singletons, None
    result = forge_embeddings.embed_texts([row_text(run, row_id) for row_id in row_ids], url=args.embeddings_url)
    if not result["ok"]:
        return singletons, f"embeddings unavailable ({result['reason']}); every row was enriched on its own"
    vectors = [forge_embeddings.normalize(vector) for vector in result["vectors"]]
    cluster_of_row = {}
    for index, component in enumerate(forge_embeddings.cluster_components(vectors, args.cluster_threshold)):
        for position in component:
            cluster_of_row[row_ids[position]] = index
    return cluster_of_row, None


def attempt_enrichment(service, messages, args):
    """Returns ``(status, value, note)``, or ``(None, None, error)``."""
    try:
        payload, _record = forge_llm.call_json_with_retry(
            service, messages, cache_prompt=args.cache_prompt, timeout=args.request_timeout, task="enrich"
        )
    except forge_llm.ChatError as error:
        return None, None, str(error)
    if not isinstance(payload, dict):
        return None, None, "response was not a JSON object"
    status = payload.get("status")
    if status == "needs_review":
        note = payload.get("note")
        if blank(note):
            return None, None, "a needs_review response requires a note"
        return "needs_review", None, str(note)
    if status != "completed":
        return None, None, f'status must be "completed" or "needs_review", received {status!r}'
    value = payload.get("value")
    if not isinstance(value, str) or blank(value):
        return None, None, "a completed response requires a nonblank string value"
    if len(value) > MAX_ENRICHMENT_CHARACTERS:
        return None, None, f"value is {len(value)} characters; one cell must be under {MAX_ENRICHMENT_CHARACTERS}"
    return "completed", value.strip(), None


def enrich_row(service, system, run, row_id, args, objection=None):
    """Enrich one row, with one corrective retry on an unusable response."""
    payload = {"row": run["rows"][str(row_id)]}
    if objection:
        # The objection rides in the user message so the system prefix stays
        # byte-identical to the one the bulk run warmed.
        payload["reviewerObjection"] = objection
    messages = [
        {"role": "system", "content": system},
        {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
    ]
    status, value, note = attempt_enrichment(service, messages, args)
    if status is None:
        repair = [*messages, {"role": "user", "content": f"That response was unusable: {note}. Return corrected JSON only."}]
        status, value, note = attempt_enrichment(service, repair, args)
        if status is None:
            raise UserError(note)
    return status, value, note


VERIFY_SYSTEM = (
    "You are reviewing a spreadsheet column filled in by a faster model without reasoning.\n"
    "For each row you get its input columns and the value written into the output column.\n"
    "Flag a row when the value is actually wrong: unsupported by the row's own data, an\n"
    "invented identifier or quantity, an answer to a different question than the column asks,\n"
    "or a value whose form is inconsistent with the other rows shown in a way that would break\n"
    "sorting or filtering. A defensible value is 'ok' even if you would have worded it\n"
    "differently; taste is not an error.\n"
    "Some rows carry 'alsoAppliedTo': the same value was written into those rows too, because\n"
    "they were judged to mean the same thing. Check that the value is right for every one of\n"
    "them, and flag the row if it is not — say which of those rows it does not fit. A value\n"
    "that is correct for the row shown but wrong for the rows it was copied to is still wrong."
)
# Enough of the reusing rows to judge the merge without inflating every packet.
VERIFY_REUSE_SAMPLE = 5


def enrichment_verification_payload(run, entry, reused_rows):
    payload = {
        "id": str(entry["rowId"]),
        "input": run["rows"][str(entry["rowId"])],
        "value": entry["value"],
    }
    if reused_rows:
        payload["reusedBy"] = len(reused_rows)
        payload["alsoAppliedTo"] = [run["rows"][str(row_id)] for row_id in reused_rows[:VERIFY_REUSE_SAMPLE]]
    return payload


def verification_sample(results, limit):
    """Rows to review: every representative whose answer was reused, then a
    spread of the others.

    Reviewing every row is not worth its cost here — enrichment is high volume
    and low variance — but a representative is not an ordinary row. Its answer
    was copied to every member of its cluster, so a wrong one is wrong many
    times over, and those are always reviewed even past the sample limit.
    """
    completed = [entry for entry in results if entry["status"] == "completed"]
    reusing_rows = {}
    for entry in completed:
        representative = entry.get("derivedFrom")
        if representative is not None:
            reusing_rows.setdefault(representative, []).append(entry["rowId"])
    representatives = [entry for entry in completed if reusing_rows.get(entry["rowId"])]
    others = [entry for entry in completed if not reusing_rows.get(entry["rowId"]) and not entry.get("derivedFrom")]
    chosen = list(representatives)
    room = max(0, limit - len(chosen))
    if others and room:
        # Evenly spaced rather than the first N, so the sample is spread across
        # the sheet, and deterministic so a rerun reviews the same rows.
        step = max(1, len(others) // room)
        chosen.extend(others[::step][:room])
    chosen.sort(key=lambda entry: entry["rowId"])
    return chosen, reusing_rows


def verify_enrichments(args, run_directory, run, system):
    """Review a weighted sample on the thinking model, and redo what it flags."""
    results = load_results(run_directory)
    sample, reusing_rows = verification_sample(results, args.verify_sample)
    completed = sum(1 for entry in results if entry["status"] == "completed")
    if not sample:
        return None
    think = forge_llm.resolve_think_or_chat(base_url=args.think_url, model=args.think_model)
    if not think["enabled"]:
        return {"skipped": "no thinking service is configured"}
    items = [enrichment_verification_payload(run, entry, reusing_rows.get(entry["rowId"], [])) for entry in sample]
    progress(f"verifying {len(items)} of {completed} enriched rows on {think['url']}")
    try:
        verdicts = forge_verify.verify_packets(
            think,
            VERIFY_SYSTEM,
            items,
            journal_path=run_directory / "verified.jsonl",
            packet_size=args.verify_packet_size,
            background=True,
            timeout=args.request_timeout,
            progress=progress,
        )
    except forge_verify.VerificationError as error:
        return {"skipped": str(error)}

    flagged = [
        (item, verdicts[item["id"]]["reason"])
        for item in items
        if verdicts.get(item["id"], {}).get("verdict") == forge_verify.VERDICT_FLAG
    ]

    def redo(item, reason):
        return enrich_row(think, system, run, int(item["id"]), args, objection=reason)

    escalations = forge_verify.escalate(flagged, redo, journal_path=run_directory / "verified.jsonl", progress=progress)
    rebuilt = 0
    unresolved = 0
    for identifier, outcome in escalations.items():
        if outcome.get("resumed"):
            continue  # committed when it was first escalated
        row_id = int(identifier)
        had_reusers = bool(reusing_rows.get(row_id))
        if outcome["ok"]:
            status, value, note = outcome["value"]
            record_row(run_directory, run, row_id, status, value, note or "re-answered with reasoning after review")
            if had_reusers:
                progress(f"[escalate] row {row_id} was reused by {len(reusing_rows[row_id])} rows; re-answering each on its own")
                added, failed = rebuild_reusing_rows(run_directory, run, think, system, row_id, reason_for(verdicts, identifier), args)
                rebuilt += added
                unresolved += failed
        else:
            record_row(run_directory, run, row_id, "needs_review", None, f"review flagged this and re-answering failed: {outcome['detail']}")
            unresolved += 1
            if had_reusers:
                unresolved += orphan_reusing_rows(run_directory, run, row_id, outcome["detail"])
    summary = forge_verify.summarize(verdicts, escalations)
    summary["sampled"] = len(items)
    summary["completed"] = completed
    summary["coverage"] = "sample plus every reused answer" if len(items) < completed else "all rows"
    if rebuilt:
        summary["reusingRowsReAnsweredIndividually"] = rebuilt
    if unresolved:
        summary["needsReview"] = summary.get("needsReview", 0) + unresolved
    return summary


def reason_for(verdicts, identifier):
    return verdicts.get(identifier, {}).get("reason", "")


def rebuild_reusing_rows(run_directory, run, think, system, representative, reason, args):
    """Re-answer the rows that copied a rejected value, one at a time.

    A reviewer who rejects a value because it does not fit the rows it was
    copied to has rejected the *merge*, not just the wording. Propagating a
    replacement value would repeat the same mistake with different words — on a
    deliberately over-broad threshold this turned ten distinct tickets into ten
    copies of one category and still reported success. So the group is broken up
    and each row is answered on its own.

    Returns ``(rebuilt, unresolved)``.
    """
    derived = [entry for entry in load_results(run_directory) if entry.get("derivedFrom") == representative]
    rebuilt = 0
    unresolved = 0
    for entry in derived[:MAX_REBUILT_ROWS]:
        row_id = entry["rowId"]
        try:
            status, value, note = enrich_row(think, system, run, row_id, args, objection=reason)
        except (UserError, OSError) as error:
            unresolved += 1
            record_row(run_directory, run, row_id, "needs_review", None, f"review rejected the answer copied from row {representative}, and re-answering failed: {error}")
            continue
        if status == "completed":
            rebuilt += 1
        else:
            unresolved += 1
        record_row(run_directory, run, row_id, status, value, note or f"re-answered on its own after review rejected the answer copied from row {representative}")
    for entry in derived[MAX_REBUILT_ROWS:]:
        unresolved += 1
        record_row(
            run_directory, run, entry["rowId"], "needs_review", None,
            f"review rejected the answer copied from row {representative}; too many rows reused it to re-answer them all automatically",
        )
    return rebuilt, unresolved


def orphan_reusing_rows(run_directory, run, representative, detail):
    """Mark rows that copied a value review rejected and could not be replaced."""
    derived = [entry for entry in load_results(run_directory) if entry.get("derivedFrom") == representative]
    for entry in derived:
        record_row(
            run_directory, run, entry["rowId"], "needs_review", None,
            f"copied from row {representative}, which review flagged and could not be re-answered: {detail}",
        )
    return len(derived)


def harmonize_trailing_period(run_directory, run):
    """Make the column agree with itself about the trailing period.

    Asking the prompt for this backfires. Naming punctuation as something to keep
    consistent makes a non-thinking model attend to it without giving it a rule:
    measured on ten rows, adding that clause took trailing periods from one cell
    to six. The script is in a better position anyway — a single stateless call
    cannot see the other rows, and this one can see all of them.

    Only a trailing period on a multi-word value is touched, so a decimal or a
    numbered value is left alone, and only toward whichever form the column
    already prefers. Returns the number of cells changed.
    """
    results = load_results(run_directory)
    completed = [entry for entry in results if entry["status"] == "completed" and isinstance(entry.get("value"), str)]
    eligible = [entry for entry in completed if len(entry["value"].split()) > 1]
    if len(eligible) < 2:
        return 0
    with_period = [entry for entry in eligible if entry["value"].rstrip().endswith(".")]
    without = [entry for entry in eligible if not entry["value"].rstrip().endswith(".")]
    if not with_period or not without:
        return 0
    minority = with_period if len(with_period) <= len(without) else without
    for entry in minority:
        text = entry["value"].rstrip()
        entry["value"] = text[:-1].rstrip() if text.endswith(".") else f"{text}."
    run_state.atomic_write_text(
        run_directory / "row_results.jsonl",
        "\n".join(json.dumps(entry, ensure_ascii=False) for entry in results) + "\n",
    )
    return len(minority)


def command_row_process(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    source = require_source(run["source"]["path"])
    if sha256(source) != run["source"]["sha256"]:
        fail("source file changed after row-init; refusing to process")
    instruction = (args.instruction or run.get("instruction") or "").strip()
    if not instruction:
        fail("no enrichment instruction; pass --instruction here or set one at row-init")
    service = forge_llm.resolve_service("chat", base_url=args.base_url, model=args.model)
    if not service["enabled"]:
        fail("connectedServices.chat is disabled; configure the local chat endpoint before processing")
    system = enrichment_system_prompt(run, instruction)

    results = load_results(run_directory)
    recorded = {entry["rowId"] for entry in results}
    pending = [row_id for row_id in run["eligibleRows"] if row_id not in recorded]
    if args.limit and args.limit > 0:
        pending = pending[: args.limit]

    # Clustered over every eligible row, not just the pending ones, so a resumed
    # run can still reuse an answer produced before the interruption.
    cluster_of_row, cluster_note = reuse_clusters(run, run["eligibleRows"], args)
    cached = {}
    for entry in results:
        if entry["status"] == "completed" and not entry.get("derivedFrom"):
            cluster = cluster_of_row.get(entry["rowId"])
            if cluster is not None:
                cached.setdefault(cluster, (entry["rowId"], entry["value"]))

    processed = 0
    reused = 0
    failures = 0
    for row_id in pending:
        cluster = cluster_of_row.get(row_id)
        hit = cached.get(cluster)
        processed += 1
        if hit is not None:
            representative, value = hit
            record_row(run_directory, run, row_id, "completed", value, f"reused the answer from row {representative}", {"derivedFrom": representative})
            reused += 1
            progress(f"[{processed}/{len(pending)}] row {row_id}: reused row {representative}")
            continue
        progress(f"[{processed}/{len(pending)}] row {row_id}")
        try:
            status, value, note = enrich_row(service, system, run, row_id, args)
        except (UserError, OSError) as error:
            failures += 1
            record_row(run_directory, run, row_id, "needs_review", None, f"enrichment failed: {error}")
            progress(f"[{processed}/{len(pending)}] row {row_id}: needs review ({error})")
            continue
        record_row(run_directory, run, row_id, status, value, note)
        if status == "completed":
            if cluster is not None:
                cached[cluster] = (row_id, value)
        else:
            failures += 1
        progress(f"[{processed}/{len(pending)}] row {row_id}: {status}")

    harmonized = harmonize_trailing_period(run_directory, run)
    verification = verify_enrichments(args, run_directory, run, system) if args.verify else None
    remaining = len(run["eligibleRows"]) - len(load_results(run_directory))
    report = {
        "processed": processed,
        "modelCalls": processed - reused,
        "reusedFromNearIdenticalRows": reused,
        "needsReview": failures,
        "remaining": remaining,
        "verification": verification,
        "nextAction": "row-process" if remaining else "row-finalize",
    }
    if harmonized:
        report["trailingPeriodHarmonized"] = harmonized
    if cluster_note:
        report["clustering"] = cluster_note
    print(json.dumps(report, ensure_ascii=False))


def result_map(results):
    return {result["rowId"]: result for result in results}


def write_delimited_output(source, output, run, results):
    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    rows = csv_rows(source)
    values = result_map(results)
    with output.open("x", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter, lineterminator="\n")
        for row_number, row in enumerate(rows, start=1):
            if row_number == 1:
                writer.writerow(row + [run["outputColumn"]])
                continue
            result = values.get(row_number)
            value = result["value"] if result and result["status"] == "completed" else ""
            writer.writerow(row + [value])


def write_xlsx_output(source, output, run, results):
    openpyxl = openpyxl_module()
    workbook = openpyxl.load_workbook(source, data_only=False, keep_links=True)
    try:
        if run["sheet"] not in workbook.sheetnames:
            fail(f"source sheet no longer exists: {run['sheet']}")
        worksheet = workbook[run["sheet"]]
        headers = [display_header(worksheet.cell(row=1, column=index).value) for index in range(1, worksheet.max_column + 1)]
        if run["outputColumn"] in headers:
            fail(f"output column now exists in source: {run['outputColumn']}")
        output_column = worksheet.max_column + 1
        source_style_column = max(1, output_column - 1)
        header = worksheet.cell(row=1, column=output_column, value=run["outputColumn"])
        source_header = worksheet.cell(row=1, column=source_style_column)
        if source_header.has_style:
            header._style = copy(source_header._style)
        if source_header.number_format:
            header.number_format = source_header.number_format
        source_letter = openpyxl.utils.get_column_letter(source_style_column)
        output_letter = openpyxl.utils.get_column_letter(output_column)
        if source_letter in worksheet.column_dimensions:
            source_dimension = worksheet.column_dimensions[source_letter]
            worksheet.column_dimensions[output_letter].width = source_dimension.width
            worksheet.column_dimensions[output_letter].hidden = source_dimension.hidden
        for row_id, result in result_map(results).items():
            if result["status"] != "completed":
                continue
            cell = worksheet.cell(row=row_id, column=output_column, value=result["value"])
            source_cell = worksheet.cell(row=row_id, column=source_style_column)
            if source_cell.has_style:
                cell._style = copy(source_cell._style)
            if source_cell.number_format:
                cell.number_format = source_cell.number_format
        temporary = output.with_name(f".{output.name}.{os.getpid()}.tmp.xlsx")
        try:
            workbook.save(temporary)
            os.link(temporary, output)
        finally:
            temporary.unlink(missing_ok=True)
    finally:
        workbook.close()


def review_report(run, results):
    lines = [
        "# Row Enrichment Review",
        "",
        f"- Source: `{run['source']['path']}`",
        f"- Sheet: `{run['sheet']}`",
        f"- Output column: `{run['outputColumn']}`",
        f"- Eligible rows: {len(run['eligibleRows'])}",
        f"- Blank rows skipped automatically: {len(run['blankRows'])}",
        "",
        "## Results",
        "",
    ]
    counts = Counter(result["status"] for result in results)
    for status in ["completed", "skipped", "failed", "needs_review"]:
        lines.append(f"- {status}: {counts[status]}")
    lines.extend(["", "## Rows Requiring Attention", ""])
    attention = [result for result in results if result["status"] != "completed"]
    if not attention:
        lines.append("None.")
    else:
        lines.extend(["| Source row | Status | Note |", "|---:|---|---|"])
        for result in attention:
            note = str(result.get("note") or "").replace("|", "\\|").replace("\n", " ")
            lines.append(f"| {result['rowId']} | {result['status']} | {note} |")
    if run["source"]["format"] == "xlsx":
        lines.extend(
            [
                "",
                "## XLSX Preservation Warning",
                "",
                "XLSX output was written through openpyxl. Common cells, formulas, sheets, and basic styles are preserved, but macros, external links, advanced charts, embedded objects, and unsupported Excel features may not survive round-tripping.",
                "",
            ]
        )
    return "\n".join(lines)


def command_row_finalize(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    results = load_results(run_directory)
    pending = next_pending(run, results)
    if pending is not None:
        fail(f"run is incomplete; next pending row is {pending}")
    source = require_source(run["source"]["path"])
    current_hash = sha256(source)
    if current_hash != run["source"]["sha256"]:
        fail("source file changed after row-init; refusing to finalize")
    extension = source.suffix.lower()
    output = run_directory / f"enriched{extension}"
    if output.exists():
        state = run_state.load_run_state(run_directory, "spreadsheet-row-enrichment")
        if state.get("status") == "complete":
            print(json.dumps({"output": str(output), "resumed": True, "complete": True}))
            return
        fail(f"output already exists without a completed run state: {output}")
    if extension == ".xlsx":
        write_xlsx_output(source, output, run, results)
    else:
        write_delimited_output(source, output, run, results)
    (run_directory / "review_report.md").write_text(review_report(run, results), encoding="utf-8")
    counts = Counter(result["status"] for result in results)
    (run_directory / "transform_log.md").write_text(
        "\n".join(
            [
                "# Transformation Log",
                "",
                f"- Finalized: {utc_now()}",
                f"- Source: `{source}`",
                f"- Source SHA-256: `{current_hash}`",
                f"- Sheet: `{run['sheet']}`",
                f"- Added column: `{run['outputColumn']}`",
                f"- Completed rows: {counts['completed']}",
                f"- Skipped rows: {counts['skipped']}",
                f"- Failed rows: {counts['failed']}",
                f"- Review-needed rows: {counts['needs_review']}",
                f"- Automatically blank rows: {len(run['blankRows'])}",
                f"- Output: `{output}`",
                "",
                "No source file was modified.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    run_state.update_run_state(
        run_directory,
        lambda state: {**state, "status": "complete", "phase": "complete", "nextAction": None},
        {"type": "run_completed", "output": str(output)},
    )
    print(json.dumps({"output": str(output), "completed": counts["completed"], "skipped": counts["skipped"], "failed": counts["failed"], "needsReview": counts["needs_review"]}))


def output_validation(run_directory, run, results, errors):
    source = Path(run["source"]["path"])
    output = run_directory / f"enriched{source.suffix.lower()}"
    if not output.exists():
        return
    values = result_map(results)
    if source.suffix.lower() in {".csv", ".tsv"}:
        rows = csv_rows(output)
        if not rows or not rows[0] or rows[0][-1] != run["outputColumn"]:
            errors.append("enriched output does not end with the configured output column")
            return
        for row_id, result in values.items():
            if row_id > len(rows):
                errors.append(f"enriched output is missing source row {row_id}")
                continue
            expected = result["value"] if result["status"] == "completed" else ""
            actual = rows[row_id - 1][-1] if rows[row_id - 1] else ""
            if actual != expected:
                errors.append(f"enriched output value differs at source row {row_id}")
        return
    openpyxl = openpyxl_module()
    workbook = openpyxl.load_workbook(output, read_only=True, data_only=False, keep_links=True)
    try:
        if run["sheet"] not in workbook.sheetnames:
            errors.append("enriched output is missing the configured sheet")
            return
        worksheet = workbook[run["sheet"]]
        headers = [display_header(worksheet.cell(row=1, column=index).value) for index in range(1, worksheet.max_column + 1)]
        if not headers or headers[-1] != run["outputColumn"]:
            errors.append("enriched output does not end with the configured output column")
            return
        output_column = len(headers)
        for row_id, result in values.items():
            expected = result["value"] if result["status"] == "completed" else None
            actual = worksheet.cell(row=row_id, column=output_column).value
            if actual != expected:
                errors.append(f"enriched output value differs at source row {row_id}")
    finally:
        workbook.close()


def command_validate(args):
    run_directory = require_run_directory(args.run_directory)
    run = load_run(run_directory)
    results = load_results(run_directory, strict=False)
    errors = []
    warnings = []
    eligible = run.get("eligibleRows", [])
    if len(eligible) != len(set(eligible)):
        errors.append("eligibleRows contains duplicates")
    if any(not isinstance(row_id, int) or row_id < 2 for row_id in eligible):
        errors.append("eligibleRows contains an invalid source row")
    seen = set()
    expected_order = []
    for result in results:
        row_id = result.get("rowId")
        expected_order.append(row_id)
        if row_id in seen:
            errors.append(f"duplicate result for row {row_id}")
        seen.add(row_id)
        if row_id not in eligible:
            errors.append(f"result references ineligible row {row_id}")
        if result.get("status") not in RESULT_STATUSES:
            errors.append(f"row {row_id} has invalid status {result.get('status')}")
        if result.get("status") == "completed" and not isinstance(result.get("value"), str):
            errors.append(f"completed row {row_id} has no string value")
        if result.get("status") != "completed" and not result.get("note"):
            errors.append(f"non-completed row {row_id} has no note")
    if expected_order != eligible[: len(expected_order)]:
        errors.append("results are not in eligible-row order")
    missing = [row_id for row_id in eligible if row_id not in seen]
    if missing:
        warnings.append(f"run is incomplete; {len(missing)} rows remain, beginning with row {missing[0]}")
    source = Path(run.get("source", {}).get("path", ""))
    if not source.is_file():
        errors.append("source file is missing")
    elif sha256(source) != run.get("source", {}).get("sha256"):
        errors.append("source file hash differs from row-init")
    output_validation(run_directory, run, results, errors)
    result = {"valid": not errors, "complete": not missing, "errors": errors, "warnings": warnings}
    print(json.dumps(result, indent=2))
    if errors:
        raise SystemExit(1)


def cluster_key_text(row, header_indexes, columns):
    parts = []
    for column in columns:
        value = json_value(row[header_indexes[column]])
        if blank(value):
            continue
        parts.append(str(value).strip())
    return " | ".join(parts)[:CLUSTER_KEY_CHARS]


def write_clusters_csv(path, members):
    fields = [
        "cluster_id",
        "group_size",
        "is_representative",
        "source_row",
        "similarity_to_representative",
        "key_text",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for member in members:
            writer.writerow(member)


def cluster_groups_markdown(run, members, multi_group_ids):
    by_group = {}
    for member in members:
        by_group.setdefault(member["cluster_id"], []).append(member)
    lines = [
        "# Candidate Groups",
        "",
        f"- Source: `{run['source']['path']}`",
        f"- Sheet: `{run['sheet']}`",
        f"- Grouped columns: {', '.join('`' + column + '`' for column in run['columns'])}",
        f"- Similarity threshold: {run['threshold']}",
        f"- Model: `{run['model']}`",
        f"- Rows grouped: {run['groupedRows']}",
        f"- Rows skipped (blank key): {len(run['blankKeyRows'])}",
        f"- Multi-row groups: {len(multi_group_ids)}",
        "",
        "Each multi-row group is a set of rows whose grouped columns are similar in "
        "meaning. These are candidates for review, not confirmed matches. Raise the "
        "threshold for tighter duplicate detection or lower it for broader topical "
        "grouping. Nothing is merged, deduplicated, or modified; decide what to do "
        "with each group yourself.",
        "",
    ]
    if not multi_group_ids:
        lines.append("No multi-row groups at this threshold. Every grouped row is on its own.")
        lines.append("")
        return "\n".join(lines)
    for cluster_id in multi_group_ids:
        group = by_group[cluster_id]
        lines.append(f"## Group {cluster_id} ({len(group)} rows)")
        lines.append("")
        lines.extend(["| Source row | Representative | Similarity | Key text |", "|---:|---|---:|---|"])
        for member in group:
            representative = "yes" if member["is_representative"] == "true" else ""
            key_cell = member["key_text"].replace("|", "\\|").replace("\n", " ")
            if len(key_cell) > 80:
                key_cell = key_cell[:77] + "..."
            lines.append(
                f"| {member['source_row']} | {representative} | "
                f"{member['similarity_to_representative']} | {key_cell} |"
            )
        lines.append("")
    return "\n".join(lines)


def command_cluster(args):
    source = require_source(args.input)
    if len(set(args.columns)) != len(args.columns):
        fail("--columns contains duplicates")
    threshold = args.threshold
    if not -1.0 <= threshold <= 1.0:
        fail("--threshold must be between -1 and 1")
    tables = load_tables(source, requested_sheet=args.sheet, all_xlsx_sheets=False)
    table = tables[0]
    rows = table["rows"]
    width = normalized_width(rows)
    if not rows or width == 0:
        fail("the selected table is empty")
    padded = pad_rows(rows, width)
    headers = [display_header(value) for value in padded[0]]
    validate_headers(headers)
    unknown = [column for column in args.columns if column not in headers]
    if unknown:
        fail(f"columns not found: {', '.join(unknown)}")
    header_indexes = {header: index for index, header in enumerate(headers)}

    grouped_rows = []
    key_texts = []
    blank_key_rows = []
    for row_number in range(2, len(padded) + 1):
        row = padded[row_number - 1]
        if all(blank(value) for value in row):
            continue
        key = cluster_key_text(row, header_indexes, args.columns)
        if not key:
            blank_key_rows.append(row_number)
            continue
        grouped_rows.append(row_number)
        key_texts.append(key)
    if not grouped_rows:
        fail("no rows had nonblank values in the selected columns")

    output = require_new_directory(args.output)
    try:
        result = forge_embeddings.embed_texts(key_texts, url=args.embeddings_url)
        if not result["ok"]:
            fail(
                "embeddings endpoint unavailable: "
                f"{result['reason']}. Set FORGE_EMBEDDINGS_URL or pass --embeddings-url; "
                "the cluster command requires embeddings."
            )
        vectors = [forge_embeddings.normalize(vector) for vector in result["vectors"]]
        components = forge_embeddings.cluster_components(vectors, threshold)

        members = []
        multi_group_ids = []
        for cluster_index, component in enumerate(
            sorted(components, key=lambda part: min(part)), start=1
        ):
            cluster_id = f"g{cluster_index}"
            representative_position = min(component, key=lambda position: grouped_rows[position])
            if len(component) > 1:
                multi_group_ids.append(cluster_id)
            for position in sorted(component, key=lambda position: grouped_rows[position]):
                similarity = forge_embeddings.cosine(vectors[position], vectors[representative_position])
                members.append(
                    {
                        "cluster_id": cluster_id,
                        "group_size": len(component),
                        "is_representative": "true" if position == representative_position else "false",
                        "source_row": grouped_rows[position],
                        "similarity_to_representative": f"{similarity:.3f}",
                        "key_text": key_texts[position],
                    }
                )

        run = {
            "schemaVersion": CLUSTER_SCHEMA_VERSION,
            "createdAt": utc_now(),
            "source": {
                "path": str(source),
                "format": source.suffix.lower().lstrip("."),
                "sha256": sha256(source),
                "sizeBytes": source.stat().st_size,
            },
            "sheet": table["name"],
            "columns": args.columns,
            "threshold": threshold,
            "model": result["model"],
            "dimensions": result["dimensions"],
            "groupedRows": len(grouped_rows),
            "blankKeyRows": blank_key_rows,
            "clusterCount": len(components),
            "multiRowGroupCount": len(multi_group_ids),
        }
        members.sort(key=lambda member: (member["cluster_id"], member["source_row"]))
        write_clusters_csv(output / "clusters.csv", members)
        (output / "cluster_groups.md").write_text(
            cluster_groups_markdown(run, members, multi_group_ids), encoding="utf-8"
        )
        (output / "cluster_run.json").write_text(
            json.dumps(run, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
    except BaseException:
        try:
            for child in output.iterdir():
                child.unlink()
            output.rmdir()
        except OSError:
            pass
        raise
    print(
        json.dumps(
            {
                "output": str(output),
                "groupedRows": len(grouped_rows),
                "blankKeyRows": len(blank_key_rows),
                "clusterCount": len(components),
                "multiRowGroupCount": len(multi_group_ids),
            }
        )
    )


def parser():
    root = argparse.ArgumentParser(description="Profile spreadsheets and manage resumable row enrichment.")
    subparsers = root.add_subparsers(dest="command", required=True)

    doctor = subparsers.add_parser("doctor", help="Report local spreadsheet capabilities.")
    doctor.add_argument("--json", action="store_true")
    doctor.set_defaults(handler=command_doctor)

    inspect = subparsers.add_parser("inspect", help="Create bounded Markdown and JSON data profiles.")
    inspect.add_argument("input")
    inspect.add_argument("--output", required=True)
    inspect.add_argument("--sheet")
    inspect.add_argument("--max-categories", type=int, default=20)
    inspect.set_defaults(handler=command_inspect)

    cluster = subparsers.add_parser(
        "cluster",
        help="Group rows by embedding-based similarity of one or more columns for fuzzy record linkage or categorization.",
    )
    cluster.add_argument("input")
    cluster.add_argument("--output", required=True)
    cluster.add_argument("--columns", nargs="+", required=True, help="One or more column headers whose combined text is embedded.")
    cluster.add_argument("--sheet")
    cluster.add_argument("--threshold", type=float, default=DEFAULT_CLUSTER_THRESHOLD)
    cluster.add_argument("--embeddings-url", help="Override the embeddings endpoint (default FORGE_EMBEDDINGS_URL or http://llms:8005/v1/embeddings).")
    cluster.set_defaults(handler=command_cluster)

    row_init = subparsers.add_parser("row-init", help="Initialize a resumable one-row-at-a-time enrichment run.")
    row_init.add_argument("input")
    row_init.add_argument("--output", required=True)
    row_init.add_argument("--column", required=True)
    row_init.add_argument("--sheet")
    row_init.add_argument("--input-columns", nargs="+")
    row_init.add_argument("--start-row", type=int)
    row_init.add_argument("--end-row", type=int)
    row_init.add_argument("--instruction", help="what the output column must contain; required by row-process")
    row_init.set_defaults(handler=command_row_init)

    row_next = subparsers.add_parser("row-next", help="Return exactly one pending row as JSON.")
    row_next.add_argument("run_directory")
    row_next.set_defaults(handler=command_row_next)

    row_status = subparsers.add_parser("row-status", help="Report durable row progress and source drift.")
    row_status.add_argument("run_directory")
    row_status.add_argument("--json", action="store_true")
    row_status.set_defaults(handler=command_row_status)

    row_retry = subparsers.add_parser("row-retry", help="Explicitly requeue failed rows.")
    row_retry.add_argument("run_directory")
    retry_group = row_retry.add_mutually_exclusive_group(required=True)
    retry_group.add_argument("--item", type=int)
    retry_group.add_argument("--all-failed", action="store_true")
    row_retry.set_defaults(handler=command_row_retry)

    row_record = subparsers.add_parser("row-record", help="Append one generated value or explicit disposition.")
    row_record.add_argument("run_directory")
    row_record.add_argument("--row-id", type=int, required=True)
    row_record.add_argument("--status", choices=sorted(RESULT_STATUSES), default="completed")
    row_record.add_argument("--value-file")
    row_record.add_argument("--note")
    row_record.set_defaults(handler=command_row_record)

    row_process = subparsers.add_parser(
        "row-process",
        help="Enrich every pending row without leaving the script, reusing one answer across near-identical rows, then have the thinking model review a weighted sample.",
    )
    row_process.add_argument("run_directory")
    row_process.add_argument("--instruction", help="overrides the instruction recorded at row-init")
    row_process.add_argument("--limit", type=int, help="stop after this many rows")
    row_process.add_argument("--base-url", help="chat service (default: connectedServices.chat)")
    row_process.add_argument("--model")
    row_process.add_argument("--think-url", help="thinking service used for review (default: connectedServices.think)")
    row_process.add_argument("--think-model")
    row_process.add_argument("--embeddings-url")
    row_process.add_argument(
        "--cluster-threshold",
        type=float,
        default=ROW_REUSE_THRESHOLD,
        help=f"cosine similarity above which one answer is reused across rows (default {ROW_REUSE_THRESHOLD})",
    )
    row_process.add_argument("--no-cluster", action="store_true", help="enrich every row on its own, without reuse")
    row_process.add_argument("--no-verify", action="store_true", help="skip the thinking-model review")
    row_process.add_argument("--verify-sample", type=int, default=DEFAULT_VERIFY_SAMPLE, help=f"rows to review beyond the reused answers, which are always reviewed (default {DEFAULT_VERIFY_SAMPLE})")
    row_process.add_argument("--verify-packet-size", type=int, default=15)
    row_process.add_argument("--no-cache-prompt", action="store_true")
    row_process.add_argument("--request-timeout", type=float, default=600)
    row_process.set_defaults(handler=command_row_process)

    row_finalize = subparsers.add_parser("row-finalize", help="Write a new enriched spreadsheet after all rows are disposed.")
    row_finalize.add_argument("run_directory")
    row_finalize.set_defaults(handler=command_row_finalize)

    validate = subparsers.add_parser("validate", help="Validate run state, provenance, and any finalized output.")
    validate.add_argument("run_directory")
    validate.set_defaults(handler=command_validate)
    return root


def main():
    args = parser().parse_args()
    if getattr(args, "max_categories", 1) < 1:
        fail("--max-categories must be positive")
    if hasattr(args, "no_cache_prompt"):
        args.cache_prompt = not args.no_cache_prompt
    if hasattr(args, "no_verify"):
        args.verify = not args.no_verify
    args.handler(args)


if __name__ == "__main__":
    main()
