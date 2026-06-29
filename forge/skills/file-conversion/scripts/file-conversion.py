#!/usr/bin/env python3

import argparse
import csv
import hashlib
import importlib.util
import json
import os
import posixpath
import re
import shutil
import stat
import subprocess
import sys
import tempfile
import unicodedata
import urllib.parse
import urllib.error
import urllib.request
import uuid
import zipfile
from collections import Counter
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree


LOW_TEXT_CHARACTERS = 40

# Heuristic thresholds for structured PDF -> Markdown extraction (PyMuPDF).
# These are deliberately conservative starting points; PDF layout varies widely,
# so detection is best-effort and every run discloses that it is heuristic.
PDF_HEADING_SIZE_RATIO = 1.35  # line size >= body_size * ratio => chapter-heading candidate
PDF_HEADING_MAX_WORDS = 12  # headings are short lines
PDF_HEADING_TOP_FRACTION = 0.33  # headings are likelier in the top third of a page
PDF_FOOTNOTE_SIZE_RATIO = 0.85  # line size <= body_size * ratio => small/footnote text
PDF_FOOTNOTE_BOTTOM_FRACTION = 0.62  # footnote region sits in the lower part of a page
PDF_SUPERSCRIPT_SIZE_RATIO = 0.80  # marker glyph is smaller than body text
PDF_SUPERSCRIPT_RISE_RATIO = 0.20  # marker baseline raised >= ratio * body_size
PDF_PARAGRAPH_GAP_RATIO = 0.60  # inter-line gap > ratio * line height => new paragraph
PDF_MULTICOLUMN_GAP_RATIO = 0.15  # horizontal gap > ratio * page width => column boundary
PDF_RUNNING_HEAD_PAGE_FRACTION = 0.5  # a top/bottom line on > this fraction of pages is chrome

MANIFEST_COLUMNS = [
    "source_path",
    "source_sha256",
    "source_format",
    "target_format",
    "status",
    "output_path",
    "warning_count",
    "error",
    "cover_path",
    "cover_sha256",
]

# Source group -> recognized extensions.
GROUP_EXTENSIONS = {
    "docx": {".docx"},
    "epub": {".epub"},
    "md": {".md", ".markdown"},
    "html": {".html", ".htm"},
    "pdf": {".pdf"},
    "csv": {".csv", ".tsv"},
    "xlsx": {".xlsx"},
    "txt": {".txt"},
}
EXTENSION_GROUP = {ext: group for group, exts in GROUP_EXTENSIONS.items() for ext in exts}

# Source group -> allowed --to targets.
ALLOWED_TARGETS = {
    "docx": {"md", "txt"},
    "epub": {"md"},
    "md": {"docx", "epub", "html", "txt"},
    "html": {"md", "txt"},
    "pdf": {"txt", "md"},
    "csv": {"xlsx"},
    "xlsx": {"csv"},
    "txt": {"txt"},
}
ALL_TARGETS = sorted({target for targets in ALLOWED_TARGETS.values() for target in targets})

PANDOC_WRITER = {"md": "gfm", "html": "html5", "docx": "docx", "epub": "epub3", "txt": "plain"}
TARGET_EXTENSION = {"md": "md", "html": "html", "docx": "docx", "epub": "epub", "txt": "txt", "csv": "csv", "xlsx": "xlsx"}
EPUB_CSS = Path(__file__).resolve().parent.parent / "references" / "epub.css"
EPUBCHECK_VERSION = "5.3.0"
EPUBCHECK_ARCHIVE_URL = f"https://github.com/w3c/epubcheck/releases/download/v{EPUBCHECK_VERSION}/epubcheck-{EPUBCHECK_VERSION}.zip"
EPUBCHECK_ARCHIVE_SHA256 = "6c07e68584b2e2ce2f89fe06e1246dfead3eb36b46b340e7d93524f29dcff6c5"


def fail(message, exit_code=1):
    print(f"Error: {message}", file=sys.stderr)
    raise SystemExit(exit_code)


def utc_now():
    from datetime import datetime, timezone

    return datetime.now(timezone.utc).isoformat()


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def openpyxl_available():
    return importlib.util.find_spec("openpyxl") is not None


def pymupdf_available():
    return importlib.util.find_spec("fitz") is not None


def pymupdf_version():
    if not pymupdf_available():
        return None
    import fitz

    version = getattr(fitz, "__version__", None)
    if version:
        return version
    bind = getattr(fitz, "VersionBind", None)
    return bind or "available"


def tool_version(command, args):
    if shutil.which(command) is None:
        return None
    try:
        result = subprocess.run([command, *args], capture_output=True, text=True, check=False)
    except OSError:
        return None
    if result.returncode != 0:
        return None
    combined = f"{result.stdout}\n{result.stderr}".strip()
    return combined.splitlines()[0].strip() if combined else "available"


def pandoc_version():
    return tool_version("pandoc", ["--version"])


def pdftotext_version():
    # pdftotext prints its banner to stderr and exits non-zero for -v; probe presence directly.
    if shutil.which("pdftotext") is None:
        return None
    try:
        result = subprocess.run(["pdftotext", "-v"], capture_output=True, text=True, check=False)
    except OSError:
        return None
    combined = f"{result.stdout}\n{result.stderr}".strip()
    return combined.splitlines()[0].strip() if combined else "available"


def java_version():
    return tool_version("java", ["--version"])


def default_tools_directory():
    agent_directory = os.environ.get("PI_CODING_AGENT_DIR") or os.environ.get("PI_FORGE_AGENT_DIR")
    if not agent_directory:
        agent_directory = str(Path(os.environ.get("PI_FORGE_HOME", Path.home() / ".local" / "share" / "pi-vault")) / "agent")
    return Path(agent_directory).expanduser().resolve() / "tools"


def managed_epubcheck_jar(tools_directory=None):
    root = Path(tools_directory).expanduser().resolve() if tools_directory else default_tools_directory()
    return root / "epubcheck" / EPUBCHECK_VERSION / "epubcheck.jar"


def epubcheck_candidate(command, source, path=None):
    try:
        result = subprocess.run([*command, "--version"], capture_output=True, text=True, check=False)
    except OSError:
        return None
    if result.returncode != 0:
        return None
    combined = f"{result.stdout}\n{result.stderr}".strip()
    version = combined.splitlines()[0].strip() if combined else "available"
    return {"command": command, "source": source, "path": str(path) if path else None, "version": version}


def resolve_epubcheck(tools_directory=None):
    explicit = os.environ.get("EPUBCHECK_JAR")
    java = java_version()
    if explicit:
        jar = Path(explicit).expanduser().resolve()
        if java is None or not jar.is_file():
            return None
        return epubcheck_candidate(["java", "-jar", str(jar)], "explicit-jar", jar)
    managed = managed_epubcheck_jar(tools_directory)
    if java is not None and managed.is_file():
        candidate = epubcheck_candidate(["java", "-jar", str(managed)], "managed", managed)
        if candidate:
            return candidate
    executable = shutil.which("epubcheck")
    if executable:
        return epubcheck_candidate([executable], "path", executable)
    return None


def safe_stem(path):
    raw = unicodedata.normalize("NFKC", path.stem).strip()
    safe = re.sub(r"[^\w.-]+", "-", raw, flags=re.UNICODE).strip("-")
    return safe or "file"


def require_new_directory(raw_path):
    path = Path(raw_path).expanduser().resolve()
    if path.exists():
        fail(f"output already exists: {path}")
    path.mkdir(parents=True)
    return path


def require_run_directory(raw_path):
    path = Path(raw_path).expanduser().resolve()
    if not path.is_dir():
        fail(f"run directory does not exist: {path}")
    if not (path / "conversion_manifest.csv").is_file():
        fail(f"conversion_manifest.csv is missing: {path}")
    return path


def iter_input_files(root):
    if root.is_file():
        yield root
        return
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        relative = path.relative_to(root)
        if any(part.startswith(".") for part in relative.parts):
            continue
        current = root
        linked = False
        for part in relative.parts:
            current = current / part
            if current.is_symlink():
                linked = True
                break
        if linked:
            continue
        yield path


def discover_inputs(raw_inputs):
    seen = set()
    files = []
    for raw in raw_inputs:
        root = Path(raw).expanduser().resolve()
        if not root.exists():
            fail(f"input does not exist: {root}")
        if root.is_symlink():
            fail(f"input is a symlink: {root}")
        for path in iter_input_files(root):
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                files.append(resolved)
    if not files:
        fail("no input files were found")
    return files


def text_warnings(value):
    warnings = []
    replacement = value.count("�")
    control = len(re.findall(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", value))
    non_whitespace = len(re.findall(r"\S", value))
    alphanumeric = len(re.findall(r"[^\W_]", value, flags=re.UNICODE))
    if replacement:
        warnings.append(f"Found {replacement} Unicode replacement characters; encoding may be damaged.")
    if control:
        warnings.append(f"Found {control} unexpected control characters.")
    if non_whitespace > 100 and alphanumeric / non_whitespace < 0.05:
        warnings.append("Output has an unusually low proportion of letters and numbers; review for garbled text.")
    if non_whitespace == 0:
        warnings.append("No readable text was produced.")
    return warnings


def pandoc_ast(source):
    if pandoc_version() is None:
        raise RuntimeError("Pandoc is required (macOS: brew install pandoc; Debian/Ubuntu: apt install pandoc)")
    result = subprocess.run(
        ["pandoc", str(source), "--from=markdown", "--to=json", f"--resource-path={source.parent}"],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"pandoc could not parse Markdown: {result.stderr.strip() or 'unknown error'}")
    return json.loads(result.stdout)


def ast_text(value):
    if isinstance(value, list):
        return "".join(ast_text(item) for item in value)
    if not isinstance(value, dict):
        return ""
    kind = value.get("t")
    content = value.get("c")
    if kind in {"Str", "Code", "Math", "MetaString"}:
        if isinstance(content, str):
            return content
        if isinstance(content, list) and content:
            return str(content[-1])
    if kind in {"Space", "SoftBreak", "LineBreak"}:
        return " "
    return ast_text(content)


def walk_ast(value):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from walk_ast(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk_ast(child)


def markdown_metadata_present(ast, key):
    return key in ast.get("meta", {})


def markdown_metadata_value(ast, key):
    return ast_text(ast.get("meta", {}).get(key)).strip()


def analyze_markdown(source, ast):
    warnings = []
    missing_images = []
    h1_count = 0
    for node in walk_ast(ast.get("blocks", [])):
        kind = node.get("t")
        content = node.get("c")
        if kind == "Header" and isinstance(content, list) and content and content[0] == 1:
            h1_count += 1
        elif kind == "Table" and isinstance(content, list):
            column_count = len(content[2]) if len(content) > 2 and isinstance(content[2], list) else 0
            table_text = ast_text(content)
            if column_count > 4:
                warnings.append(f"Table has {column_count} columns and may be difficult to read on a small screen.")
            if any(len(token) > 40 for token in re.findall(r"\S+", table_text)):
                warnings.append("Table contains an unbreakable value longer than 40 characters; review it on narrow screens.")
        elif kind in {"RawBlock", "RawInline"} and isinstance(content, list) and content and content[0] in {"html", "html4", "html5"}:
            raw = str(content[-1])
            if re.search(r"\b(?:rowspan|colspan)\s*=", raw, flags=re.IGNORECASE):
                warnings.append("Raw HTML uses spanning table cells; support varies across EPUB readers.")
            else:
                warnings.append("Raw HTML may not render consistently across EPUB readers.")
        elif kind == "Image" and isinstance(content, list) and len(content) >= 3:
            alt = ast_text(content[1]).strip()
            target = content[2][0] if isinstance(content[2], list) and content[2] else ""
            if not alt:
                warnings.append(f"Image has no alternative text: {target or '(unknown image)'}")
            parsed = urllib.parse.urlparse(target)
            if parsed.scheme in {"http", "https"}:
                warnings.append(f"Remote image must be fetched during conversion and may reduce reproducibility: {target}")
            elif parsed.scheme != "data":
                local_target = urllib.parse.unquote(parsed.path)
                image_path = Path(local_target) if Path(local_target).is_absolute() else source.parent / local_target
                if local_target and not image_path.is_file():
                    missing_images.append(str(image_path.resolve()))
    if h1_count == 0:
        warnings.append("Markdown has no level-one heading; EPUB contains one readable content section without chapter-level navigation.")
    if missing_images:
        raise RuntimeError(f"referenced local image is missing: {missing_images[0]}")
    return warnings


def jpeg_dimensions_and_progressive(data):
    if not data.startswith(b"\xff\xd8"):
        return None
    index = 2
    while index + 4 <= len(data):
        if data[index] != 0xFF:
            index += 1
            continue
        marker = data[index + 1]
        index += 2
        if marker in {0xD8, 0xD9}:
            continue
        if index + 2 > len(data):
            break
        length = int.from_bytes(data[index : index + 2], "big")
        if length < 2 or index + length > len(data):
            break
        if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
            if length < 7:
                break
            height = int.from_bytes(data[index + 3 : index + 5], "big")
            width = int.from_bytes(data[index + 5 : index + 7], "big")
            return width, height, marker == 0xC2
        index += length
    return None


def validate_cover(path):
    data = path.read_bytes()
    width = height = None
    if data.startswith(b"\x89PNG\r\n\x1a\n") and len(data) >= 24:
        width = int.from_bytes(data[16:20], "big")
        height = int.from_bytes(data[20:24], "big")
    else:
        jpeg = jpeg_dimensions_and_progressive(data)
        if jpeg is None:
            raise RuntimeError("cover must be a baseline JPEG or PNG image")
        width, height, progressive = jpeg
        if progressive:
            raise RuntimeError("progressive JPEG covers are not supported by all target readers; use a baseline JPEG or PNG")
    warnings = []
    if width > 800 or height > 1200 or len(data) > 1024 * 1024:
        warnings.append(
            f"Cover is {width}x{height} and {len(data)} bytes; large covers can index slowly on the XTEINK X4."
        )
    return warnings


def run_epub(source, output, options):
    ast = pandoc_ast(source)
    warnings = analyze_markdown(source, ast)
    command = [
        "pandoc",
        str(source),
        "--from=markdown",
        "--to=epub3",
        "--toc",
        "--toc-depth=1",
        "--split-level=1",
        f"--resource-path={source.parent}",
        f"--css={EPUB_CSS}",
        f"--metadata=identifier:urn:uuid:{uuid.uuid5(uuid.NAMESPACE_URL, sha256(source))}",
        "--output",
        str(output),
    ]
    metadata = {
        "title": options.get("title"),
        "author": options.get("author"),
        "lang": options.get("language"),
        "date": options.get("date"),
    }
    if not metadata["title"] and not markdown_metadata_present(ast, "title"):
        metadata["title"] = source.stem
    if not metadata["lang"]:
        if not markdown_metadata_present(ast, "lang"):
            metadata["lang"] = markdown_metadata_value(ast, "language") or "en-US"
    for key, value in metadata.items():
        if value:
            command.append(f"--metadata={key}:{value}")
    cover = options.get("cover")
    if cover is not None:
        warnings.extend(validate_cover(cover))
        command.append(f"--epub-cover-image={cover}")
    result = subprocess.run(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"pandoc failed: {result.stderr.strip() or 'unknown error'}")
    if result.stderr.strip():
        warnings.append(f"Pandoc reported: {result.stderr.strip()}")
    return warnings


def epub_structure(path):
    try:
        archive = zipfile.ZipFile(path)
    except (OSError, zipfile.BadZipFile) as error:
        raise RuntimeError(f"EPUB archive is invalid: {error}")
    with archive:
        names = set(archive.namelist())
        if "META-INF/rights.xml" in names:
            raise RuntimeError("EPUB appears to be DRM-protected and cannot be converted")
        if "META-INF/encryption.xml" in names:
            try:
                encryption = ElementTree.fromstring(archive.read("META-INF/encryption.xml"))
            except ElementTree.ParseError as error:
                raise RuntimeError(f"EPUB encryption metadata is malformed: {error}")
            allowed = {
                "http://www.idpf.org/2008/embedding",
                "http://ns.adobe.com/pdf/enc#RC",
            }
            algorithms = {
                element.get("Algorithm")
                for element in encryption.iter()
                if element.tag.endswith("EncryptionMethod") and element.get("Algorithm")
            }
            if any(algorithm not in allowed for algorithm in algorithms):
                raise RuntimeError("EPUB contains encrypted resources that cannot be converted")
        if "META-INF/container.xml" not in names:
            raise RuntimeError("EPUB is missing META-INF/container.xml")
        try:
            container = ElementTree.fromstring(archive.read("META-INF/container.xml"))
        except ElementTree.ParseError as error:
            raise RuntimeError(f"EPUB container metadata is malformed: {error}")
        rootfile = next((element for element in container.iter() if element.tag.endswith("rootfile")), None)
        package_path = rootfile.get("full-path") if rootfile is not None else None
        if not package_path or package_path not in names:
            raise RuntimeError("EPUB container does not reference an existing package document")
        try:
            package = ElementTree.fromstring(archive.read(package_path))
        except ElementTree.ParseError as error:
            raise RuntimeError(f"EPUB package metadata is malformed: {error}")
        package_dir = posixpath.dirname(package_path)
        manifest = {}
        nav_path = None
        cover_path = None
        scripted = False
        unsupported_media = set()
        for item in package.iter():
            if not item.tag.endswith("item") or not item.get("id"):
                continue
            target = local_archive_target(package_dir, item.get("href", ""))
            manifest[item.get("id")] = target
            properties = item.get("properties", "").split()
            media_type = item.get("media-type", "")
            if "nav" in properties:
                nav_path = target
            if "cover-image" in properties:
                cover_path = target
            if "scripted" in properties:
                scripted = True
            if media_type.startswith(("audio/", "video/")):
                unsupported_media.add(media_type)
        spine = []
        for itemref in package.iter():
            if itemref.tag.endswith("itemref") and itemref.get("idref") in manifest:
                target = manifest[itemref.get("idref")]
                if target:
                    spine.append(target)
        fixed_layout = any(
            element.tag.endswith("meta")
            and element.get("property") == "rendition:layout"
            and "pre-paginated" in "".join(element.itertext())
            for element in package.iter()
        )
        labels = {}
        if nav_path and nav_path in names:
            try:
                nav = ElementTree.fromstring(archive.read(nav_path))
            except ElementTree.ParseError:
                nav = None
            if nav is not None:
                nav_dir = posixpath.dirname(nav_path)
                for navigation in nav.iter():
                    epub_type = navigation.get("{http://www.idpf.org/2007/ops}type", "")
                    if not navigation.tag.endswith("nav") or "toc" not in epub_type.split():
                        continue
                    for link in navigation.iter():
                        if not link.tag.endswith("a") or not link.get("href"):
                            continue
                        target = local_archive_target(nav_dir, link.get("href"))
                        label = " ".join("".join(link.itertext()).split())
                        if target and label:
                            labels.setdefault(target, label)
                    break
        return {
            "spine": spine,
            "labels": labels,
            "coverPath": cover_path,
            "fixedLayout": fixed_layout,
            "scripted": scripted,
            "unsupportedMedia": sorted(unsupported_media),
        }


def empty_span_marker(block):
    if block.get("t") != "Para" or not isinstance(block.get("c"), list) or len(block["c"]) != 1:
        return None
    span = block["c"][0]
    if span.get("t") != "Span" or not isinstance(span.get("c"), list) or len(span["c"]) != 2:
        return None
    attributes, content = span["c"]
    if content or not isinstance(attributes, list) or not attributes:
        return None
    identifier = attributes[0]
    return identifier if isinstance(identifier, str) and identifier.endswith((".xhtml", ".html", ".htm")) else None


def markdown_slug(value):
    normalized = unicodedata.normalize("NFKC", value).casefold()
    normalized = re.sub(r"[^\w\s-]", "", normalized, flags=re.UNICODE)
    return re.sub(r"[-\s]+", "-", normalized).strip("-") or "chapter"


def text_inlines(value):
    words = value.split()
    inlines = []
    for index, word in enumerate(words):
        if index:
            inlines.append({"t": "Space"})
        inlines.append({"t": "Str", "c": word})
    return inlines


def transform_ast_list(items):
    transformed = []
    skip_landmarks = False
    for item in items:
        if isinstance(item, dict) and item.get("t") == "RawBlock":
            content = item.get("c", [])
            raw = str(content[-1]) if isinstance(content, list) and content else ""
            if re.search(r"<nav\b[^>]*\bepub:type=[\"']landmarks[\"']", raw, flags=re.IGNORECASE):
                skip_landmarks = True
                continue
            if skip_landmarks and re.search(r"</nav\s*>", raw, flags=re.IGNORECASE):
                skip_landmarks = False
                continue
        if skip_landmarks:
            continue
        if isinstance(item, dict) and item.get("t") == "Div":
            content = item.get("c", [])
            attributes = content[0] if isinstance(content, list) and content else []
            classes = attributes[1] if isinstance(attributes, list) and len(attributes) > 1 else []
            if "section" in classes and len(content) > 1:
                transformed.extend(transform_ast_list(content[1]))
                continue
        if isinstance(item, dict) and item.get("t") == "Figure":
            content = item.get("c", [])
            if isinstance(content, list) and len(content) > 2:
                transformed.extend(transform_ast_list(content[2]))
                continue
        transformed.append(transform_ast_value(item))
    return transformed


def transform_ast_value(value):
    if isinstance(value, list):
        return transform_ast_list(value)
    if isinstance(value, dict):
        return {key: transform_ast_value(child) for key, child in value.items()}
    return value


def chapter_label(marker, structure):
    basename = posixpath.basename(marker)
    spine_targets = [target for target in structure["spine"] if posixpath.basename(target) == basename]
    for target in spine_targets:
        if target in structure["labels"]:
            return structure["labels"][target], target
    return None, spine_targets[0] if len(spine_targets) == 1 else None


def collect_section_aliases(ast):
    aliases = {}
    for node in walk_ast(ast.get("blocks", [])):
        if node.get("t") != "Div" or not isinstance(node.get("c"), list) or len(node["c"]) < 2:
            continue
        attributes, blocks = node["c"][:2]
        identifier = attributes[0] if isinstance(attributes, list) and attributes else ""
        if not identifier:
            continue
        header = next((item for item in walk_ast(blocks) if item.get("t") == "Header"), None)
        if header:
            aliases[identifier] = markdown_slug(ast_text(header.get("c", [None, None, []])[-1]))
    return aliases


def rewrite_epub_targets(ast, output, media_directory, aliases, chapter_aliases):
    media_root = media_directory.resolve()
    for node in walk_ast(ast):
        if node.get("t") not in {"Image", "Link"} or not isinstance(node.get("c"), list) or len(node["c"]) < 3:
            continue
        target = node["c"][2]
        if not isinstance(target, list) or not target:
            continue
        raw_target = target[0]
        if node.get("t") == "Image":
            image = Path(raw_target)
            if image.is_absolute() and image.resolve().is_relative_to(media_root):
                target[0] = Path(os.path.relpath(image, output.parent)).as_posix()
            continue
        parsed = urllib.parse.urlparse(raw_target)
        fragment = urllib.parse.unquote(parsed.fragment)
        if fragment in aliases:
            target[0] = f"#{aliases[fragment]}"
            continue
        link_path = urllib.parse.unquote(parsed.path)
        for chapter_path, slug in chapter_aliases.items():
            if link_path and posixpath.basename(link_path) == posixpath.basename(chapter_path):
                target[0] = f"#{slug}"
                break


def clean_epub_ast(ast, structure):
    aliases = collect_section_aliases(ast)
    segments = []
    marker = None
    blocks = []
    for block in ast.get("blocks", []):
        next_marker = empty_span_marker(block)
        if next_marker:
            segments.append((marker, blocks))
            marker = next_marker
            blocks = []
        else:
            blocks.append(block)
    segments.append((marker, blocks))
    output_blocks = []
    warnings = []
    chapter_aliases = {}
    for marker, segment in segments:
        if marker and posixpath.basename(marker) in {"nav.xhtml", "title_page.xhtml"}:
            continue
        cleaned = transform_ast_list(segment)
        label, chapter_path = chapter_label(marker or "", structure)
        h1 = next(
            (
                node
                for node in walk_ast(cleaned)
                if node.get("t") == "Header" and isinstance(node.get("c"), list) and node["c"] and node["c"][0] == 1
            ),
            None,
        )
        if label and h1 is None:
            cleaned.insert(0, {"t": "Header", "c": [1, ["", [], []], text_inlines(label)]})
            warnings.append(f'Synthesized level-one chapter heading from EPUB navigation: "{label}".')
            h1 = cleaned[0]
        if chapter_path and h1:
            slug = markdown_slug(ast_text(h1.get("c", [None, None, []])[-1]))
            chapter_aliases[chapter_path] = slug
            aliases[posixpath.basename(chapter_path)] = slug
        output_blocks.extend(cleaned)
    ast["blocks"] = output_blocks
    return aliases, chapter_aliases, warnings


def copy_epub_cover(source, structure, media_directory):
    cover_path = structure.get("coverPath")
    if not cover_path:
        return None
    with zipfile.ZipFile(source) as archive:
        if cover_path not in archive.namelist():
            return None
        suffix = PurePosixPath(cover_path).suffix.lower() or ".img"
        destination = media_directory / f"cover{suffix}"
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_bytes(archive.read(cover_path))
        return destination


def run_epub_to_markdown(source, output, converted_directory):
    if pandoc_version() is None:
        raise RuntimeError("Pandoc is required (macOS: brew install pandoc; Debian/Ubuntu: apt install pandoc)")
    structure = epub_structure(source)
    media_directory = converted_directory / "media" / safe_stem(source)
    warnings = ["EPUB styling, typography, pagination, and reading-system behavior are not represented in Markdown."]
    if structure["fixedLayout"]:
        warnings.append("The source uses fixed layout; its page composition cannot be preserved in reflowable Markdown.")
    if structure["scripted"]:
        warnings.append("The source contains scripts; interactive behavior was removed from Markdown.")
    if structure["unsupportedMedia"]:
        warnings.append(f"Audio or video resources were not represented in Markdown: {', '.join(structure['unsupportedMedia'])}.")
    with tempfile.TemporaryDirectory(prefix="pi-forge-epub-") as temporary:
        temporary_directory = Path(temporary)
        raw_ast_path = temporary_directory / "raw.json"
        command = [
            "pandoc",
            str(source),
            "--from=epub",
            "--to=json",
            f"--extract-media={media_directory}",
            "--output",
            str(raw_ast_path),
        ]
        result = subprocess.run(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            raise RuntimeError(f"pandoc failed to read EPUB: {result.stderr.strip() or 'unknown error'}")
        ast = json.loads(raw_ast_path.read_text(encoding="utf-8"))
        aliases, chapter_aliases, cleanup_warnings = clean_epub_ast(ast, structure)
        warnings.extend(cleanup_warnings)
        rewrite_epub_targets(ast, output, media_directory, aliases, chapter_aliases)
        transformed_path = temporary_directory / "clean.json"
        transformed_path.write_text(json.dumps(ast), encoding="utf-8")
        result = subprocess.run(
            [
                "pandoc",
                str(transformed_path),
                "--from=json",
                "--to=gfm",
                "--standalone",
                "--wrap=none",
                "--output",
                str(output),
            ],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(f"pandoc failed to write Markdown: {result.stderr.strip() or 'unknown error'}")
    if media_directory.is_dir() and any(media_directory.rglob("*")):
        warnings.append(f"Embedded media was extracted to {media_directory}; referenced paths are relative to the Markdown file.")
    cover = copy_epub_cover(source, structure, media_directory)
    if cover:
        warnings.append(f"The EPUB cover was preserved at {cover} and was not inserted into the Markdown body.")
    warnings.extend(text_warnings(output.read_text(encoding="utf-8", errors="replace")))
    return warnings


def run_pandoc(source, output, target, media_dir):
    if pandoc_version() is None:
        raise RuntimeError("Pandoc is required (macOS: brew install pandoc; Debian/Ubuntu: apt install pandoc)")
    command = ["pandoc", str(source), "-t", PANDOC_WRITER[target], "-o", str(output)]
    if target == "html":
        command.append("--standalone")
    warnings = []
    if media_dir is not None:
        command.append(f"--extract-media={media_dir}")
    result = subprocess.run(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(f"pandoc failed: {result.stderr.strip() or 'unknown error'}")
    if media_dir is not None and media_dir.is_dir() and any(media_dir.rglob("*")):
        warnings.append(f"Embedded media was extracted to {media_dir}; it is referenced from the converted file.")
    if target in {"md", "txt"} and output.is_file():
        warnings.extend(text_warnings(output.read_text(encoding="utf-8", errors="replace")))
    if target in {"docx", "html"}:
        warnings.append("Conversion preserves common structure; complex layout, styles, and footnotes may not survive.")
    return warnings


def run_pdftotext(source, output, target):
    if pdftotext_version() is None:
        raise RuntimeError("pdftotext (Poppler) is required (macOS: brew install poppler; Debian/Ubuntu: apt install poppler-utils)")
    result = subprocess.run(
        ["pdftotext", "-layout", "-enc", "UTF-8", str(source), "-"],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"pdftotext failed: {result.stderr.strip() or 'unknown error'}")
    raw = result.stdout.replace("\r\n", "\n")
    warnings = []
    if target == "md":
        pages = [page.strip("\n") for page in raw.split("\f")]
        body = "\n\n".join(page for page in pages if page)
        text = body + "\n" if body else ""
        warnings.append("PDF Markdown is unstructured extracted text; headings and tables are not reconstructed.")
    else:
        text = raw if raw.endswith("\n") or raw == "" else raw + "\n"
    output.write_text(text, encoding="utf-8")
    if len(re.findall(r"[^\W_]", text, flags=re.UNICODE)) < LOW_TEXT_CHARACTERS:
        warnings.append("Very little text was extracted; the PDF may be scanned. Use document-ingest for OCR.")
    warnings.extend(text_warnings(text))
    return warnings


def pdf_norm_chrome(text):
    # Normalize a line for running-head/footer detection: drop digits so page
    # numbers cluster, lowercase, and collapse whitespace.
    stripped = re.sub(r"\d+", "", text)
    return re.sub(r"\s+", " ", stripped).strip().lower()


def pdf_collect_lines(doc):
    # Flatten the document into an ordered list of line records carrying the
    # geometry and spans needed for heuristic classification, plus the modal
    # body font size (character-weighted so prose dominates over headings).
    lines = []
    size_counter = Counter()
    for page_index, page in enumerate(doc):
        page_dict = page.get_text("dict", sort=True)
        page_height = page_dict.get("height") or page.rect.height or 1.0
        page_width = page_dict.get("width") or page.rect.width or 1.0
        page_lines = []
        for block in page_dict.get("blocks", []):
            if block.get("type") != 0:
                continue
            block_number = block.get("number", 0)
            for line in block.get("lines", []):
                spans = [span for span in line.get("spans", []) if span.get("text", "")]
                if not spans:
                    continue
                text = "".join(span["text"] for span in spans)
                if not text.strip():
                    continue
                bbox = line.get("bbox", (0, 0, 0, 0))
                max_size = max(round(span["size"], 1) for span in spans)
                dominant = max(spans, key=lambda span: len(span["text"].strip()))
                for span in spans:
                    size_counter[round(span["size"])] += len(span["text"].strip())
                page_lines.append(
                    {
                        "page": page_index,
                        "page_height": page_height,
                        "page_width": page_width,
                        "block": (page_index, block_number),
                        "x0": bbox[0],
                        "x1": bbox[2],
                        "top": bbox[1],
                        "bottom": bbox[3],
                        "height": max(bbox[3] - bbox[1], 1.0),
                        "max_size": max_size,
                        "base_y": dominant.get("origin", (0, bbox[3]))[1],
                        "text": text,
                        "spans": spans,
                        "is_page_top": False,
                        "is_page_bottom": False,
                    }
                )
        if page_lines:
            ordered = sorted(page_lines, key=lambda rec: (rec["top"], rec["x0"]))
            ordered[0]["is_page_top"] = True
            ordered[-1]["is_page_bottom"] = True
            lines.extend(page_lines)
    body_size = float(size_counter.most_common(1)[0][0]) if size_counter else 0.0
    return lines, body_size


def pdf_running_heads(lines, page_count):
    if page_count < 3:
        return set()
    counter = Counter()
    for line in lines:
        if line["is_page_top"] or line["is_page_bottom"]:
            key = pdf_norm_chrome(line["text"])
            if key:
                counter[key] += 1
    threshold = page_count * PDF_RUNNING_HEAD_PAGE_FRACTION
    return {key for key, count in counter.items() if count > threshold}


def pdf_is_page_number(text):
    return re.fullmatch(r"[\divxlcdmIVXLCDM.\s\-–—]+", text.strip()) is not None


def pdf_body_line_text(line, body_size, markers):
    # Rebuild the visible text of a body line, replacing superscript reference
    # markers with sentinels (\x00M{idx}\x00) resolved after note matching.
    base_y = line["base_y"]
    parts = []
    for span in line["spans"]:
        stripped = span["text"].strip()
        raised = base_y - span.get("origin", (0, base_y))[1]
        if (
            re.fullmatch(r"\d{1,3}", stripped)
            and span["size"] <= body_size * PDF_SUPERSCRIPT_SIZE_RATIO
            and raised >= PDF_SUPERSCRIPT_RISE_RATIO * body_size
        ):
            idx = len(markers)
            markers.append(int(stripped))
            parts.append(f"\x00M{idx}\x00")
        else:
            parts.append(span["text"])
    return "".join(parts)


def pdf_paragraphs(body_lines, body_size, markers):
    paragraphs = []
    current = ""
    prev = None
    for line in body_lines:
        text = pdf_body_line_text(line, body_size, markers).strip()
        if not text:
            prev = line
            continue
        if prev is None:
            new_para = True
        elif line["page"] != prev["page"]:
            # Paragraphs spanning a page break are split; joining across pages
            # would merge unrelated text given running heads are already dropped.
            new_para = True
        else:
            # Block boundaries are unreliable (many PDFs emit one block per
            # visual line), so paragraph breaks are detected by vertical gap.
            gap = line["top"] - prev["bottom"]
            new_para = gap > PDF_PARAGRAPH_GAP_RATIO * line["height"]
        if new_para:
            if current.strip():
                paragraphs.append(current.strip())
            current = text
        elif current.endswith("-") and len(current) >= 2 and current[-2].isalpha() and text[:1].islower():
            current = current[:-1] + text
        else:
            current = f"{current} {text}"
        prev = line
    if current.strip():
        paragraphs.append(current.strip())
    return paragraphs


def pdf_match_notes(markers, footnotes, chapter_index):
    # Match in-text markers to footnote bodies by printed number (FIFO per
    # number to tolerate per-page numbering restarts). Returns the inline
    # replacement per marker index, ordered (id, text) definitions, and counts.
    from collections import deque

    positions = {}
    for index, (number, _body) in enumerate(footnotes):
        positions.setdefault(number, deque()).append(index)
    consumed = [False] * len(footnotes)
    replacements = {}
    definitions = []
    sequence = 0
    unmatched_markers = 0
    for marker_index, number in enumerate(markers):
        queue = positions.get(number)
        if queue:
            note_index = queue.popleft()
            consumed[note_index] = True
            sequence += 1
            footnote_id = f"c{chapter_index}-{sequence}"
            replacements[marker_index] = f"[^{footnote_id}]"
            definitions.append((footnote_id, footnotes[note_index][1]))
        else:
            replacements[marker_index] = str(number)
            unmatched_markers += 1
    orphans = 0
    for index, (_number, body) in enumerate(footnotes):
        if not consumed[index]:
            sequence += 1
            definitions.append((f"c{chapter_index}-{sequence}", body))
            orphans += 1
    return replacements, definitions, unmatched_markers, orphans


def pdf_render_chapter(chapter, chapter_index, body_size, stats):
    heading = chapter["heading"] or stats["default_title"]
    markers = []
    paragraphs = pdf_paragraphs(chapter["body_lines"], body_size, markers)
    replacements, definitions, unmatched, orphans = pdf_match_notes(markers, chapter["footnotes"], chapter_index)
    stats["unmatched_markers"] += unmatched
    stats["orphan_notes"] += orphans
    body = "\n\n".join(paragraphs)
    for marker_index, replacement in replacements.items():
        body = body.replace(f"\x00M{marker_index}\x00", replacement)
    out = [f"# {heading}", ""]
    if body:
        out.append(body)
        out.append("")
    if definitions:
        out.append("## Endnotes")
        out.append("")
        for footnote_id, note_text in definitions:
            clean = re.sub(r"\s+", " ", note_text).strip()
            out.append(f"[^{footnote_id}]: {clean}")
        out.append("")
    return "\n".join(out).rstrip("\n")


def run_pdf_to_markdown_structured(source, output):
    if importlib.util.find_spec("fitz") is None:
        raise RuntimeError(
            "PyMuPDF (fitz) is required for PDF->Markdown (pip install pymupdf); PDF->txt still works via pdftotext"
        )
    import fitz

    try:
        doc = fitz.open(source)
    except Exception as error:  # fitz raises library-specific errors on malformed input
        raise RuntimeError(f"PDF could not be opened: {error}")
    try:
        if doc.is_encrypted and doc.authenticate("") == 0:
            raise RuntimeError("PDF is encrypted or password-protected and cannot be converted")
        page_count = doc.page_count
        lines, body_size = pdf_collect_lines(doc)
    finally:
        doc.close()

    warnings = [
        "PDF Markdown structure (headings, footnotes) is reconstructed heuristically and may be incomplete or misattributed.",
        "Tables, figures, and complex layout are not reconstructed.",
    ]

    if body_size <= 0 or not lines:
        text = ""
        output.write_text(text, encoding="utf-8")
        warnings.append("Very little text was extracted; the PDF may be scanned. Use document-ingest for OCR.")
        warnings.extend(text_warnings(text))
        return warnings

    drop_chrome = pdf_running_heads(lines, page_count)

    chapters = []
    current = None
    headings_detected = 0
    leading_synthesized = False
    open_footnote = None  # (page, chapter id) tracking for continuation lines
    open_footnote_page = None

    def ensure_chapter():
        nonlocal current, leading_synthesized
        if current is None:
            current = {"heading": None, "body_lines": [], "footnotes": []}
            chapters.append(current)
            leading_synthesized = True
        return current

    for line in lines:
        text = line["text"].strip()
        if not text:
            continue
        # Drop running heads/footers and standalone page numbers.
        if (line["is_page_top"] or line["is_page_bottom"]) and pdf_norm_chrome(text) in drop_chrome:
            continue
        if (line["is_page_top"] or line["is_page_bottom"]) and pdf_is_page_number(text):
            continue

        top_fraction = line["top"] / line["page_height"]
        is_heading = (
            line["max_size"] >= body_size * PDF_HEADING_SIZE_RATIO
            and len(text.split()) <= PDF_HEADING_MAX_WORDS
            and not pdf_is_page_number(text)
        )
        in_footnote_region = (
            top_fraction >= PDF_FOOTNOTE_BOTTOM_FRACTION and line["max_size"] <= body_size * PDF_FOOTNOTE_SIZE_RATIO
        )

        if is_heading:
            current = {"heading": text, "body_lines": [], "footnotes": []}
            chapters.append(current)
            headings_detected += 1
            open_footnote = None
            open_footnote_page = None
            continue

        if in_footnote_region:
            chapter = ensure_chapter()
            match = re.match(r"^\s*(\d{1,3})[.)\]\s]\s*(.*)$", text)
            if match:
                chapter["footnotes"].append((int(match.group(1)), match.group(2).strip()))
                open_footnote = len(chapter["footnotes"]) - 1
                open_footnote_page = line["page"]
                continue
            if open_footnote is not None and open_footnote_page == line["page"] and chapter["footnotes"]:
                number, body = chapter["footnotes"][open_footnote]
                chapter["footnotes"][open_footnote] = (number, f"{body} {text}".strip())
                continue
            # Small bottom text with no number and no open note: treat as body.

        ensure_chapter()["body_lines"].append(line)

    if not chapters:
        text = ""
        output.write_text(text, encoding="utf-8")
        warnings.append("Very little text was extracted; the PDF may be scanned. Use document-ingest for OCR.")
        warnings.extend(text_warnings(text))
        return warnings

    stats = {"default_title": safe_stem(source), "unmatched_markers": 0, "orphan_notes": 0}
    rendered = [pdf_render_chapter(chapter, index + 1, body_size, stats) for index, chapter in enumerate(chapters)]
    text = "\n\n".join(part for part in rendered if part)
    text = text + "\n" if text else ""
    output.write_text(text, encoding="utf-8")

    if headings_detected == 0:
        warnings.append("No chapter headings were detected; the document was emitted as a single synthesized chapter.")
    elif leading_synthesized:
        warnings.append("Content before the first detected heading was placed under a synthesized chapter title.")
    if stats["unmatched_markers"]:
        warnings.append(
            f"{stats['unmatched_markers']} footnote markers had no matching footnote text and were left inline."
        )
    if stats["orphan_notes"]:
        warnings.append(
            f"{stats['orphan_notes']} footnotes had no in-text marker and were appended to Endnotes unlinked."
        )
    if pdf_has_columns(lines):
        warnings.append("Multiple text columns were detected; the reading order may be incorrect.")
    if len(re.findall(r"[^\W_]", text, flags=re.UNICODE)) < LOW_TEXT_CHARACTERS:
        warnings.append("Very little text was extracted; the PDF may be scanned. Use document-ingest for OCR.")
    warnings.extend(text_warnings(text))
    return warnings


def pdf_has_columns(lines):
    body_lines = [line for line in lines if not line["is_page_top"] and not line["is_page_bottom"]]
    if len(body_lines) < 10:
        return False
    left = sum(1 for line in body_lines if line["x1"] < line["page_width"] * 0.5)
    right = sum(1 for line in body_lines if line["x0"] > line["page_width"] * 0.55)
    total = len(body_lines)
    return left > total * 0.2 and right > total * 0.2


def csv_rows(path):
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.reader(handle, delimiter=delimiter))
    except UnicodeDecodeError:
        raise RuntimeError(f"input is not valid UTF-8: {path}")


def run_csv_to_xlsx(source, output):
    if not openpyxl_available():
        raise RuntimeError("openpyxl is required to write XLSX; install it for the active Python 3 environment")
    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row in csv_rows(source):
        worksheet.append(row)
    workbook.save(output)
    return ["Values were written as text; numeric and date typing was not inferred."]


def sheet_filename(title, used):
    cleaned = re.sub(r"[^\w.-]+", "-", unicodedata.normalize("NFKC", title), flags=re.UNICODE).strip("-") or "sheet"
    candidate = cleaned
    index = 1
    while candidate.lower() in used:
        index += 1
        candidate = f"{cleaned}-{index}"
    used.add(candidate.lower())
    return candidate


def run_xlsx_to_csv(source, converted_dir, stem):
    if not openpyxl_available():
        raise RuntimeError("openpyxl is required to read XLSX; install it for the active Python 3 environment")
    import openpyxl

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    warnings = []
    try:
        worksheets = list(workbook.worksheets)
        if not worksheets:
            raise RuntimeError("workbook has no sheets")
        if len(worksheets) == 1:
            outputs = [converted_dir / f"{stem}.csv"]
            targets = [(worksheets[0], outputs[0])]
        else:
            sheet_dir = converted_dir / stem
            sheet_dir.mkdir(parents=True, exist_ok=True)
            used = set()
            targets = [(ws, sheet_dir / f"{sheet_filename(ws.title, used)}.csv") for ws in worksheets]
            outputs = [path for _, path in targets]
            warnings.append(f"Workbook had {len(worksheets)} sheets; each was written to a separate CSV under {sheet_dir}.")
        for worksheet, out_path in targets:
            with out_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle, lineterminator="\n")
                for row in worksheet.iter_rows(values_only=True):
                    writer.writerow(["" if cell is None else cell for cell in row])
    finally:
        workbook.close()
    warnings.append("Formulas were exported as their last-computed values; macros, charts, and styling were dropped.")
    return outputs[0], warnings


def run_txt_cleanup(source, output):
    try:
        raw = source.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        raise RuntimeError(f"input is not valid UTF-8: {source}")
    lines = [line.rstrip() for line in raw.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    text = re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip("\n")
    text = text + "\n" if text else ""
    output.write_text(text, encoding="utf-8")
    return text_warnings(text)


def unique_output(converted_dir, stem, extension, used):
    candidate = f"{stem}.{extension}"
    index = 1
    while candidate.lower() in used:
        index += 1
        candidate = f"{stem}-{index}.{extension}"
    used.add(candidate.lower())
    return converted_dir / candidate


def convert_one(source, target, converted_dir, used_names, options=None):
    options = options or {}
    group = EXTENSION_GROUP.get(source.suffix.lower())
    if group is None or target not in ALLOWED_TARGETS.get(group, set()):
        return {"status": "skipped", "output": None, "warnings": [], "error": f"{source.suffix or '(none)'} cannot convert to {target}"}
    stem = safe_stem(source)
    try:
        if group == "csv":
            output = unique_output(converted_dir, stem, "xlsx", used_names)
            warnings = run_csv_to_xlsx(source, output)
        elif group == "xlsx":
            output, warnings = run_xlsx_to_csv(source, converted_dir, stem)
        elif group == "pdf":
            output = unique_output(converted_dir, stem, TARGET_EXTENSION[target], used_names)
            if target == "md":
                warnings = run_pdf_to_markdown_structured(source, output)
            else:
                warnings = run_pdftotext(source, output, target)
        elif group == "txt":
            output = unique_output(converted_dir, stem, "txt", used_names)
            warnings = run_txt_cleanup(source, output)
        elif group == "epub":
            output = unique_output(converted_dir, stem, "md", used_names)
            warnings = run_epub_to_markdown(source, output, converted_dir)
        elif group == "md" and target == "epub":
            output = unique_output(converted_dir, stem, "epub", used_names)
            warnings = run_epub(source, output, options)
        else:  # docx, md, html via pandoc
            output = unique_output(converted_dir, stem, TARGET_EXTENSION[target], used_names)
            media_dir = (converted_dir / "media" / stem) if group in {"docx", "html"} else None
            warnings = run_pandoc(source, output, target, media_dir)
    except Exception as error:  # handlers may raise tool/library-specific errors; keep the batch going
        message = str(error) or error.__class__.__name__
        return {"status": "failed", "output": None, "warnings": [], "error": message}
    status = "needs_review" if warnings else "success"
    return {"status": status, "output": output, "warnings": warnings, "error": ""}


def command_doctor(args):
    xlsx = openpyxl_available()
    openpyxl_ver = None
    if xlsx:
        import openpyxl

        openpyxl_ver = openpyxl.__version__
    pandoc = pandoc_version()
    pdftotext = pdftotext_version()
    pymupdf = pymupdf_version()
    java = java_version()
    epubcheck = resolve_epubcheck()
    managed_jar = managed_epubcheck_jar()
    install_command = f"python3 {Path(__file__).resolve()} install-epubcheck"
    result = {
        "python": sys.version.split()[0],
        "pandoc": pandoc is not None,
        "pandocVersion": pandoc,
        "pdftotext": pdftotext is not None,
        "pdftotextVersion": pdftotext,
        "pymupdf": pymupdf is not None,
        "pymupdfVersion": pymupdf,
        "java": java is not None,
        "javaVersion": java,
        "epubcheck": epubcheck is not None,
        "epubcheckVersion": epubcheck["version"] if epubcheck else None,
        "epubcheckSource": epubcheck["source"] if epubcheck else None,
        "epubcheckPath": epubcheck["path"] if epubcheck else None,
        "managedEpubcheckPath": str(managed_jar),
        "epubcheckInstallCommand": install_command,
        "xlsx": xlsx,
        "openpyxlVersion": openpyxl_ver,
        "remediation": [],
    }
    if pandoc is None:
        result["remediation"].append("Install Pandoc for DOCX/Markdown/HTML conversions (macOS: brew install pandoc; Debian/Ubuntu: apt install pandoc).")
    if pdftotext is None:
        result["remediation"].append("Install Poppler for PDF->txt conversion (macOS: brew install poppler; Debian/Ubuntu: apt install poppler-utils).")
    if pymupdf is None:
        result["remediation"].append("Install PyMuPDF for structured PDF->Markdown (pip install pymupdf); PDF->txt still works with Poppler.")
    if not xlsx:
        result["remediation"].append("Install openpyxl for CSV<->XLSX conversion in the active Python 3 environment.")
    if java is None:
        result["remediation"].append(
            "Install a Java runtime before managed EPUBCheck (macOS: brew install openjdk; Debian/Ubuntu: apt install default-jre)."
        )
    if epubcheck is None:
        result["remediation"].append(
            f"Install pinned EPUBCheck {EPUBCHECK_VERSION} with: {install_command}. Built-in structural validation remains available."
        )
    if args.json:
        print(json.dumps(result, indent=2))
        return
    print(f"Python: {result['python']}")
    print(f"Pandoc (DOCX/MD/HTML/EPUB): {pandoc or 'unavailable'}")
    print(f"pdftotext (PDF->txt): {pdftotext or 'unavailable'}")
    print(f"PyMuPDF (PDF->Markdown structure): {pymupdf or 'unavailable'}")
    print(f"Java (managed EPUBCheck runtime): {java or 'unavailable'}")
    print(f"EPUBCheck (optional EPUB conformance): {epubcheck['version'] if epubcheck else 'unavailable'}")
    if epubcheck:
        print(f"EPUBCheck source: {epubcheck['source']} ({epubcheck['path'] or 'command on PATH'})")
    print(f"openpyxl (CSV<->XLSX): {'available ' + openpyxl_ver if xlsx else 'unavailable'}")
    for action in result["remediation"]:
        print(f"Action: {action}")


def safe_epubcheck_archive(archive):
    expected_root = f"epubcheck-{EPUBCHECK_VERSION}"
    for info in archive.infolist():
        path = PurePosixPath(info.filename)
        if "\\" in info.filename or path.is_absolute() or ".." in path.parts or not path.parts or path.parts[0] != expected_root:
            raise RuntimeError(f"EPUBCheck archive contains an unsafe or unexpected path: {info.filename}")
        mode = info.external_attr >> 16
        if stat.S_ISLNK(mode):
            raise RuntimeError(f"EPUBCheck archive contains a symbolic link: {info.filename}")


def command_install_epubcheck(args):
    java = java_version()
    if java is None:
        fail(
            "a working Java runtime is required for EPUBCheck "
            "(macOS: brew install openjdk; Debian/Ubuntu: apt install default-jre)"
        )
    if args.expected_sha256 and not args.archive:
        fail("--expected-sha256 can only be used with --archive")
    expected_sha256 = (args.expected_sha256 or EPUBCHECK_ARCHIVE_SHA256).lower()
    if not re.fullmatch(r"[0-9a-f]{64}", expected_sha256):
        fail("expected EPUBCheck SHA-256 must contain exactly 64 hexadecimal characters")
    tools_directory = Path(args.tools_directory).expanduser().resolve() if args.tools_directory else default_tools_directory()
    target = managed_epubcheck_jar(tools_directory).parent
    if target.exists():
        existing = resolve_epubcheck(tools_directory)
        if existing and existing["source"] == "managed":
            print(
                json.dumps(
                    {
                        "installed": False,
                        "alreadyInstalled": True,
                        "version": existing["version"],
                        "path": str(target),
                    }
                )
            )
            return
        fail(f"managed EPUBCheck directory already exists but is not usable: {target}")
    target.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="epubcheck-install-", dir=target.parent) as temporary:
        temporary_directory = Path(temporary)
        archive_path = temporary_directory / f"epubcheck-{EPUBCHECK_VERSION}.zip"
        if args.archive:
            source_archive = Path(args.archive).expanduser().resolve()
            if not source_archive.is_file():
                fail(f"EPUBCheck archive does not exist: {source_archive}")
            shutil.copyfile(source_archive, archive_path)
        else:
            try:
                request = urllib.request.Request(EPUBCHECK_ARCHIVE_URL, headers={"User-Agent": "pi-forge/epubcheck-installer"})
                with urllib.request.urlopen(request, timeout=60) as response:
                    with archive_path.open("wb") as handle:
                        shutil.copyfileobj(response, handle)
            except (OSError, urllib.error.URLError) as error:
                fail(f"could not download EPUBCheck {EPUBCHECK_VERSION}: {error}")
        actual_sha256 = sha256(archive_path)
        if actual_sha256 != expected_sha256:
            fail(f"EPUBCheck archive SHA-256 mismatch: expected {expected_sha256}, got {actual_sha256}")
        try:
            with zipfile.ZipFile(archive_path) as archive:
                safe_epubcheck_archive(archive)
                archive.extractall(temporary_directory / "extracted")
        except zipfile.BadZipFile as error:
            fail(f"EPUBCheck archive is invalid: {error}")
        extracted = temporary_directory / "extracted" / f"epubcheck-{EPUBCHECK_VERSION}"
        jar = extracted / "epubcheck.jar"
        if not jar.is_file() or not (extracted / "lib").is_dir():
            fail("EPUBCheck archive is missing epubcheck.jar or its lib directory")
        candidate = epubcheck_candidate(["java", "-jar", str(jar)], "managed", jar)
        if not candidate or EPUBCHECK_VERSION not in candidate["version"]:
            fail(f"downloaded EPUBCheck did not report expected version {EPUBCHECK_VERSION}")
        os.replace(extracted, target)
    print(
        json.dumps(
            {
                "installed": True,
                "alreadyInstalled": False,
                "version": candidate["version"],
                "path": str(target),
                "archiveSha256": actual_sha256,
            }
        )
    )


def command_convert(args):
    if args.target not in ALL_TARGETS:
        fail(f"unsupported target: {args.target}; expected one of {', '.join(ALL_TARGETS)}")
    from_extension = None
    if args.from_extension:
        from_extension = args.from_extension if args.from_extension.startswith(".") else f".{args.from_extension}"
        from_extension = from_extension.lower()
        if from_extension not in EXTENSION_GROUP:
            fail(f"--from extension is not a recognized source: {from_extension}")
    files = discover_inputs(args.inputs)
    selected_files = [path for path in files if not from_extension or path.suffix.lower() == from_extension]
    book_options_used = any([args.cover, args.title, args.author, args.language, args.date])
    if book_options_used and args.target != "epub":
        fail("--cover and EPUB metadata options can only be used with --to epub")
    eligible_markdown = [path for path in selected_files if EXTENSION_GROUP.get(path.suffix.lower()) == "md"]
    if book_options_used and len(eligible_markdown) != 1:
        fail("--cover and EPUB metadata options require exactly one Markdown source")
    cover = None
    if args.cover:
        cover = Path(args.cover).expanduser().resolve()
        if not cover.is_file():
            fail(f"cover image does not exist: {cover}")
        try:
            validate_cover(cover)
        except RuntimeError as error:
            fail(str(error))
    output = require_new_directory(args.output)
    converted_dir = output / "converted"
    converted_dir.mkdir()
    manifest_path = output / "conversion_manifest.csv"
    used_names = set()
    rows = []
    counts = Counter()
    for source in files:
        if from_extension and source.suffix.lower() != from_extension:
            continue
        options = {
            "cover": cover,
            "title": args.title,
            "author": args.author,
            "language": args.language,
            "date": args.date,
        }
        outcome = convert_one(source, args.target, converted_dir, used_names, options)
        counts[outcome["status"]] += 1
        output_rel = str(outcome["output"].relative_to(output)) if outcome["output"] else ""
        rows.append(
            [
                str(source),
                sha256(source),
                source.suffix.lower().lstrip(".") or "(none)",
                args.target,
                outcome["status"],
                output_rel,
                len(outcome["warnings"]),
                outcome["error"],
                str(cover) if cover else "",
                sha256(cover) if cover else "",
            ]
        )
        append_log(output, "conversion_log.md", [f"- {utc_now()} {outcome['status'].upper()} {source} -> {output_rel or '(none)'}"])
        if outcome["warnings"]:
            append_log(output, "warnings.md", [f"## {source.name} -> {args.target}", "", *[f"- {warning}" for warning in outcome["warnings"]], ""])
    if not rows:
        fail("no files matched the requested conversion; nothing was written")
    with manifest_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(MANIFEST_COLUMNS)
        writer.writerows(rows)
    print(
        json.dumps(
            {
                "runDirectory": str(output),
                "target": args.target,
                "success": counts["success"],
                "needsReview": counts["needs_review"],
                "skipped": counts["skipped"],
                "failed": counts["failed"],
            }
        )
    )


def append_log(run_directory, name, lines):
    path = run_directory / name
    existing = path.read_text(encoding="utf-8") if path.is_file() else ""
    path.write_text(existing + "\n".join(lines) + "\n", encoding="utf-8")


def local_archive_target(base, href):
    parsed = urllib.parse.urlparse(href)
    if parsed.scheme or parsed.netloc:
        return None
    path = urllib.parse.unquote(parsed.path)
    if not path:
        return base
    normalized = posixpath.normpath(posixpath.join(base, path))
    if normalized == ".." or normalized.startswith("../"):
        return None
    return normalized


def epubcheck_report_messages(value):
    messages = []
    if isinstance(value, dict):
        severity = str(value.get("severity", "")).upper()
        if severity in {"FATAL", "ERROR", "WARNING", "USAGE"}:
            code = value.get("ID") or value.get("id") or value.get("code") or "EPUBCheck"
            text = value.get("message") or value.get("text") or value.get("description") or "validation finding"
            messages.append((severity, f"{code}: {text}"))
        for child in value.values():
            messages.extend(epubcheck_report_messages(child))
    elif isinstance(value, list):
        for child in value:
            messages.extend(epubcheck_report_messages(child))
    return messages


def run_epubcheck_validation(path):
    resolved = resolve_epubcheck()
    if resolved is None:
        return [], ["EPUBCheck is unavailable; only built-in structural validation was performed."]
    with tempfile.TemporaryDirectory(prefix="pi-forge-epubcheck-") as temporary:
        report_path = Path(temporary) / "report.json"
        result = subprocess.run(
            [*resolved["command"], str(path), "--json", str(report_path)],
            capture_output=True,
            text=True,
            check=False,
        )
        messages = []
        if report_path.is_file():
            try:
                messages = epubcheck_report_messages(json.loads(report_path.read_text(encoding="utf-8")))
            except (json.JSONDecodeError, OSError):
                messages = []
        errors = [message for severity, message in messages if severity in {"FATAL", "ERROR"}]
        warnings = [message for severity, message in messages if severity in {"WARNING", "USAGE"}]
        if result.returncode != 0 and not errors:
            details = (result.stdout + "\n" + result.stderr).strip()
            errors.append(f"EPUBCheck failed: {details or f'exit status {result.returncode}'}")
        return errors, warnings


def validate_epub(path, require_cover=False):
    errors = []
    warnings = []
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            if not names or names[0] != "mimetype":
                errors.append("EPUB mimetype must be the first archive entry.")
            elif archive.getinfo("mimetype").compress_type != zipfile.ZIP_STORED:
                errors.append("EPUB mimetype entry must be stored without compression.")
            elif archive.read("mimetype") != b"application/epub+zip":
                errors.append("EPUB mimetype entry has invalid content.")
            if "META-INF/container.xml" not in names:
                errors.append("EPUB is missing META-INF/container.xml.")
                return errors, warnings
            try:
                container = ElementTree.fromstring(archive.read("META-INF/container.xml"))
            except ElementTree.ParseError as error:
                errors.append(f"EPUB container.xml is not well-formed XML: {error}")
                return errors, warnings
            rootfile = container.find("{urn:oasis:names:tc:opendocument:xmlns:container}rootfiles/{urn:oasis:names:tc:opendocument:xmlns:container}rootfile")
            package_path = rootfile.get("full-path") if rootfile is not None else None
            if not package_path or package_path not in names:
                errors.append("EPUB container does not reference an existing package document.")
                return errors, warnings
            try:
                package = ElementTree.fromstring(archive.read(package_path))
            except ElementTree.ParseError as error:
                errors.append(f"EPUB package document is not well-formed XML: {error}")
                return errors, warnings
            namespace = package.tag.split("}", 1)[0].lstrip("{") if package.tag.startswith("{") else ""
            ns = {"opf": namespace}
            package_dir = str(Path(package_path).parent)
            package_dir = "" if package_dir == "." else package_dir
            manifest = {}
            nav_path = None
            cover_path = None
            for item in package.findall("opf:manifest/opf:item", ns):
                identifier = item.get("id")
                href = item.get("href", "")
                target = local_archive_target(package_dir, href)
                if identifier:
                    manifest[identifier] = target
                if target is None or target not in names:
                    errors.append(f"EPUB manifest resource is missing or unsafe: {href or '(empty href)'}")
                if "nav" in item.get("properties", "").split():
                    nav_path = target
                if "cover-image" in item.get("properties", "").split():
                    cover_path = target
            spine_ids = [item.get("idref") for item in package.findall("opf:spine/opf:itemref", ns)]
            if not spine_ids or any(identifier not in manifest for identifier in spine_ids):
                errors.append("EPUB spine is empty or references resources outside the manifest.")
            if not nav_path or nav_path not in names:
                errors.append("EPUB manifest does not contain a navigation document.")
            if require_cover and (not cover_path or cover_path not in names):
                errors.append("EPUB does not declare the selected cover with the cover-image property.")
            for target in {value for value in manifest.values() if value and value.endswith((".xhtml", ".html", ".htm"))}:
                try:
                    ElementTree.fromstring(archive.read(target))
                except ElementTree.ParseError as error:
                    errors.append(f"EPUB content document is not well-formed XML: {target}: {error}")
            if nav_path and nav_path in names:
                try:
                    nav = ElementTree.fromstring(archive.read(nav_path))
                except ElementTree.ParseError:
                    nav = None
                nav_dir = str(Path(nav_path).parent)
                nav_dir = "" if nav_dir == "." else nav_dir
                toc_links = []
                if nav is not None:
                    for element in nav.iter():
                        if element.tag.endswith("nav") and "toc" in element.get("{http://www.idpf.org/2007/ops}type", "").split():
                            toc_links.extend(child.get("href") for child in element.iter() if child.tag.endswith("a") and child.get("href"))
                for href in toc_links:
                    target = local_archive_target(nav_dir, href)
                    if target not in names:
                        errors.append(f"EPUB table of contents link target is missing: {href}")
                    else:
                        fragment = urllib.parse.urlparse(href).fragment
                        if fragment:
                            document = ElementTree.fromstring(archive.read(target))
                            identifiers = {element.get("id") for element in document.iter() if element.get("id")}
                            if urllib.parse.unquote(fragment) not in identifiers:
                                errors.append(f"EPUB table of contents fragment is missing: {href}")
                if not toc_links:
                    warnings.append("EPUB table of contents has no chapter links; the source may have no level-one headings.")
    except (OSError, zipfile.BadZipFile, KeyError) as error:
        errors.append(f"EPUB archive is invalid: {error}")
        return errors, warnings
    epubcheck_errors, epubcheck_warnings = run_epubcheck_validation(path)
    errors.extend(f"EPUBCheck: {error}" for error in epubcheck_errors)
    warnings.extend(f"EPUBCheck: {warning}" for warning in epubcheck_warnings)
    return errors, warnings


def validate_epub_markdown(source, output):
    errors = []
    warnings = []
    try:
        ast = pandoc_ast(output)
        analyze_markdown(output, ast)
    except (RuntimeError, json.JSONDecodeError) as error:
        errors.append(f"generated Markdown could not be validated: {error}")
        return errors, warnings
    structure = epub_structure(source)
    expected_chapters = {
        path for path in structure["spine"] if path in structure["labels"]
    }
    actual_chapters = sum(
        1
        for node in walk_ast(ast.get("blocks", []))
        if node.get("t") == "Header" and isinstance(node.get("c"), list) and node["c"] and node["c"][0] == 1
    )
    if actual_chapters < len(expected_chapters):
        errors.append(
            f"generated Markdown has {actual_chapters} level-one headings for {len(expected_chapters)} EPUB chapters"
        )
    text = output.read_text(encoding="utf-8", errors="replace")
    if re.search(r"epub:type=[\"']landmarks|class=[\"'][^\"']*\bsection\b", text, flags=re.IGNORECASE):
        errors.append("generated Markdown still contains EPUB navigation or section wrapper markup")
    cover_path = structure.get("coverPath")
    if cover_path:
        suffix = PurePosixPath(cover_path).suffix.lower() or ".img"
        expected_cover = output.parent / "media" / safe_stem(source) / f"cover{suffix}"
        if not expected_cover.is_file():
            errors.append(f"extracted EPUB cover is missing: {expected_cover}")
    return errors, warnings


def command_validate(args):
    run_directory = require_run_directory(args.run_directory)
    manifest_path = run_directory / "conversion_manifest.csv"
    with manifest_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    errors = []
    warnings = []
    counts = Counter()
    for row in rows:
        counts[row.get("status")] += 1
        source = Path(row.get("source_path", ""))
        if not source.is_file():
            errors.append(f"source file is missing: {source}")
        elif sha256(source) != row.get("source_sha256"):
            errors.append(f"source file hash differs from conversion: {source}")
        cover_path = row.get("cover_path", "")
        if cover_path:
            cover = Path(cover_path)
            if not cover.is_file():
                errors.append(f"cover image is missing: {cover}")
            elif sha256(cover) != row.get("cover_sha256"):
                errors.append(f"cover image hash differs from conversion: {cover}")
        if row.get("status") in {"success", "needs_review"}:
            output = run_directory / row.get("output_path", "")
            if not row.get("output_path") or not output.exists():
                errors.append(f"converted output is missing: {row.get('output_path') or '(none)'}")
            elif row.get("target_format") == "epub":
                epub_errors, epub_warnings = validate_epub(output, require_cover=bool(cover_path))
                errors.extend(f"{output.name}: {error}" for error in epub_errors)
                warnings.extend(f"{output.name}: {warning}" for warning in epub_warnings)
            elif row.get("source_format") == "epub" and row.get("target_format") == "md":
                markdown_errors, markdown_warnings = validate_epub_markdown(source, output)
                errors.extend(f"{output.name}: {error}" for error in markdown_errors)
                warnings.extend(f"{output.name}: {warning}" for warning in markdown_warnings)
    if counts["failed"]:
        warnings.append(f"{counts['failed']} files failed conversion; see conversion_manifest.csv error column.")
    if counts["skipped"]:
        warnings.append(f"{counts['skipped']} files were skipped as unsupported for the requested target.")
    result = {
        "valid": not errors,
        "counts": {status: counts[status] for status in ["success", "needs_review", "skipped", "failed"]},
        "errors": errors,
        "warnings": warnings,
    }
    print(json.dumps(result, indent=2))
    if errors:
        raise SystemExit(1)


def parser():
    root = argparse.ArgumentParser(description="Convert files between formats while preserving originals.")
    subparsers = root.add_subparsers(dest="command", required=True)

    doctor = subparsers.add_parser("doctor", help="Report local conversion tool availability.")
    doctor.add_argument("--json", action="store_true")
    doctor.set_defaults(handler=command_doctor)

    install_epubcheck = subparsers.add_parser(
        "install-epubcheck", help=f"Install pinned EPUBCheck {EPUBCHECK_VERSION} into isolated pi-forge state."
    )
    install_epubcheck.add_argument("--tools-dir", dest="tools_directory", help="Override the managed tools root directory.")
    install_epubcheck.add_argument("--archive", help="Use a local EPUBCheck ZIP instead of downloading the official release.")
    install_epubcheck.add_argument(
        "--expected-sha256", help="Override the expected archive SHA-256 for an explicitly supplied archive."
    )
    install_epubcheck.set_defaults(handler=command_install_epubcheck)

    convert = subparsers.add_parser("convert", help="Convert files or folders to a target format.")
    convert.add_argument("inputs", nargs="+")
    convert.add_argument("--to", dest="target", choices=ALL_TARGETS, required=True)
    convert.add_argument("--output", required=True)
    convert.add_argument("--from", dest="from_extension")
    convert.add_argument("--cover", help="Baseline JPEG or PNG cover image for a single EPUB source.")
    convert.add_argument("--title", help="Override EPUB title metadata for a single source.")
    convert.add_argument("--author", help="Override EPUB author metadata for a single source.")
    convert.add_argument("--language", help="Override EPUB language metadata for a single source (for example, en-US).")
    convert.add_argument("--date", help="Override EPUB date metadata for a single source (ISO 8601 recommended).")
    convert.set_defaults(handler=command_convert)

    validate = subparsers.add_parser("validate", help="Validate a conversion run against its manifest.")
    validate.add_argument("run_directory")
    validate.set_defaults(handler=command_validate)
    return root


def main():
    args = parser().parse_args()
    args.handler(args)


if __name__ == "__main__":
    main()
